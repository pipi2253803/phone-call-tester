#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Android电话自动拨打测试程序
功能：
1. 通过USB连接检测Android手机
2. 显示手机名称和SIM卡状态
3. 自动拨打电话并控制通话时长
4. 循环拨打指定次数
5. 右侧显示详细日志
"""

import sys
import os
import subprocess
import re
import time
import threading
import platform
from datetime import datetime
from dataclasses import dataclass
from typing import Optional, List, Callable

# Windows 下隐藏控制台窗口
if platform.system() == 'Windows':
    import ctypes
    ctypes.windll.kernel32.SetConsoleCP(65001)
    ctypes.windll.kernel32.SetConsoleOutputCP(65001)

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QSpinBox, QTextEdit, QGroupBox,
    QFormLayout, QSplitter, QStatusBar, QMessageBox, QComboBox,
    QFrame, QProgressBar, QGraphicsDropShadowEffect, QTabWidget,
    QScrollArea, QGridLayout, QTableWidget, QTableWidgetItem
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer, QSize, QUrl
from PyQt6.QtGui import QFont, QColor, QTextCharFormat, QBrush, QIcon, QPalette
from PyQt6.QtWebEngineWidgets import QWebEngineView

import folium
from folium.plugins import MarkerCluster

# 现代化配色方案
# Windows 风格配色方案 - 更中性的色调
COLORS = {
    'primary': '#0078D4',        # Windows 蓝
    'primary_dark': '#005A9E',   # 深蓝色
    'primary_light': '#E5F1FB',  # 浅蓝背景
    'success': '#107C10',        # Windows 绿
    'warning': '#FFB900',        # Windows 黄
    'error': '#D83B01',          # Windows 红
    'info': '#0078D4',           # Windows 蓝
    'background': '#F3F3F3',     # Windows 背景灰
    'card_bg': '#FFFFFF',        # 纯白卡片
    'text_primary': '#323130',   # 主要文字 - 深灰
    'text_secondary': '#605E5C', # 次要文字 - 中灰
    'border': '#E1DFDD',         # 边框灰
    'divider': '#EDEBE9',        # 分隔线
}

# Windows 风格图标定义 - 使用文字符号替代 emoji
ICONS = {
    'phone': '[电话]',
    'call': '[拨号]',
    'settings': '[设置]',
    'refresh': '[刷新]',
    'start': '[开始]',
    'stop': '[停止]',
    'save': '[保存]',
    'clear': '[清空]',
    'log': '[日志]',
    'device': '[设备]',
    'sim': '[SIM]',
    'signal': '[信号]',
    'time': '[时间]',
    'count': '[次数]',
    'target': '[目标]',
    'check': '[OK]',
    'cross': '[X]',
    'warning': '[!]',
    'info': '[i]',
    'success': '[v]',
    'error': '[x]',
    'connected': '[已连接]',
    'disconnected': '[未连接]',
    'pending': '[等待]',
    # 新增图标
    'script': '[脚本]',
    'result': '[结果]',
    'map': '[地图]',
    'location': '[位置]',
    'run': '[执行]',
    'chart': '[图表]',
    'list': '[列表]',
    'calendar': '[日期]',
    'clock': '[时间]',
    'wifi': '[网络]',
    'network': '[网络]',
    'resize': '[调整]',
    'panel': '[面板]',
}

# Windows 原生字体设置
WINDOWS_FONTS = {
    'ui': 'Segoe UI',           # 主要 UI 字体
    'ui_zh': 'Microsoft YaHei UI',  # 中文 UI 字体
    'mono': 'Consolas',         # 等宽字体
    'mono_zh': 'Microsoft YaHei Mono',  # 中文等宽
}

def get_font_family() -> str:
    """获取适合当前平台的字体"""
    if platform.system() == 'Windows':
        return f"{WINDOWS_FONTS['ui']}, {WINDOWS_FONTS['ui_zh']}"
    return "Segoe UI, Microsoft YaHei, PingFang SC, sans-serif"

def get_mono_font() -> str:
    """获取适合当前平台的等宽字体"""
    if platform.system() == 'Windows':
        return f"{WINDOWS_FONTS['mono']}, {WINDOWS_FONTS['ui_zh']}"
    return "Consolas, Microsoft YaHei, monospace"


@dataclass
class SimInfo:
    """SIM卡信息数据类"""
    slot: int  # 0=卡一, 1=卡二
    state: str = "Unknown"  # 状态：就绪/无SIM卡等
    operator: str = "Unknown"  # 运营商
    phone_number: str = "Unknown"  # 手机号
    signal_level: int = 0  # 信号强度 0-4
    network_type: str = "Unknown"  # 网络类型：5G/4G/3G等


@dataclass
class DeviceInfo:
    """设备信息数据类"""
    serial: str
    name: str = "Unknown"
    model: str = "Unknown"
    android_version: str = "Unknown"  # Android版本
    sim_state: str = "Unknown"  # 主卡状态（兼容旧版本）
    sim_operator: str = "Unknown"  # 主卡运营商
    phone_number: str = "Unknown"  # 主卡号码
    sim1: SimInfo = None  # 卡一信息
    sim2: SimInfo = None  # 卡二信息
    
    def __post_init__(self):
        if self.sim1 is None:
            self.sim1 = SimInfo(slot=0)
        if self.sim2 is None:
            self.sim2 = SimInfo(slot=1)


class ADBHelper:
    """ADB工具类 - 带缓存优化和智能路径查找"""
    
    _adb_available: Optional[bool] = None
    _adb_path: Optional[str] = None
    _cache: dict = {}
    _cache_time: float = 0
    _cache_ttl: float = 5.0  # 缓存5秒
    
    @classmethod
    def _get_cache_key(cls, cmd: tuple) -> str:
        return '|'.join(cmd)
    
    @classmethod
    def _get_cached_result(cls, cmd: tuple) -> Optional[tuple]:
        """获取缓存结果"""
        if time.time() - cls._cache_time > cls._cache_ttl:
            cls._cache.clear()
            return None
        key = cls._get_cache_key(cmd)
        return cls._cache.get(key)
    
    @classmethod
    def _set_cached_result(cls, cmd: tuple, result: tuple):
        """设置缓存结果"""
        cls._cache_time = time.time()
        key = cls._get_cache_key(cmd)
        cls._cache[key] = result
    
    @classmethod
    def _find_adb_path(cls) -> Optional[str]:
        """智能查找ADB路径"""
        # 首先尝试系统PATH中的adb
        try:
            startupinfo = None
            if platform.system() == 'Windows':
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                startupinfo.wShowWindow = subprocess.SW_HIDE
            
            result = subprocess.run(
                ['adb', 'version'],
                capture_output=True,
                text=True,
                timeout=3,
                startupinfo=startupinfo
            )
            if result.returncode == 0:
                return 'adb'  # 在PATH中直接使用
        except:
            pass
        
        # Windows常见安装路径
        if platform.system() == 'Windows':
            common_paths = [
                # Android Studio 默认安装路径
                os.path.expandvars(r'%LOCALAPPDATA%\Android\Sdk\platform-tools\adb.exe'),
                os.path.expandvars(r'%USERPROFILE%\AppData\Local\Android\Sdk\platform-tools\adb.exe'),
                # 其他常见路径
                r'C:\Android\platform-tools\adb.exe',
                r'C:\Program Files\Android\platform-tools\adb.exe',
                r'C:\ProgramData\Android\platform-tools\adb.exe',
                r'C:\adb\adb.exe',
                r'C:\platform-tools\adb.exe',
                # 用户目录下
                os.path.expandvars(r'%USERPROFILE%\Downloads\platform-tools\adb.exe'),
                os.path.expandvars(r'%USERPROFILE%\Desktop\platform-tools\adb.exe'),
            ]
        else:
            # macOS/Linux
            common_paths = [
                '/usr/local/bin/adb',
                '/usr/bin/adb',
                '/opt/android-sdk/platform-tools/adb',
                os.path.expanduser('~/Library/Android/sdk/platform-tools/adb'),
                os.path.expanduser('~/Android/Sdk/platform-tools/adb'),
            ]
        
        for path in common_paths:
            if os.path.isfile(path):
                # 验证这个adb是否可用
                try:
                    # Windows 下隐藏控制台窗口
                    startupinfo = None
                    if platform.system() == 'Windows':
                        startupinfo = subprocess.STARTUPINFO()
                        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                        startupinfo.wShowWindow = subprocess.SW_HIDE
                    
                    result = subprocess.run(
                        [path, 'version'],
                        capture_output=True,
                        text=True,
                        timeout=3,
                        startupinfo=startupinfo
                    )
                    if result.returncode == 0:
                        return path
                except:
                    continue
        
        return None
    
    @classmethod
    def set_adb_path(cls, path: str) -> bool:
        """手动设置ADB路径"""
        if os.path.isfile(path):
            cls._adb_path = path
            cls._adb_available = True
            return True
        return False
    
    @classmethod
    def get_adb_path(cls) -> str:
        """获取ADB路径"""
        if cls._adb_path is None:
            cls._adb_path = cls._find_adb_path()
        return cls._adb_path or 'adb'
    
    @classmethod
    def check_adb_installed(cls) -> bool:
        """检查ADB是否已安装"""
        if cls._adb_available is not None:
            return cls._adb_available
        
        cls._adb_path = cls._find_adb_path()
        cls._adb_available = cls._adb_path is not None
        return cls._adb_available
    
    @classmethod
    def execute_command(cls, cmd: List[str], timeout: int = 30, use_cache: bool = False) -> tuple:
        """执行ADB命令 - 支持缓存和自动路径查找"""
        # 确保使用正确的ADB路径
        if len(cmd) > 0 and cmd[0] == 'adb':
            adb_path = cls.get_adb_path()
            cmd = [adb_path] + cmd[1:]
        
        cmd_tuple = tuple(cmd)
        
        # 检查缓存
        if use_cache:
            cached = cls._get_cached_result(cmd_tuple)
            if cached is not None:
                return cached
        
        try:
            # Windows 下隐藏控制台窗口
            startupinfo = None
            if platform.system() == 'Windows':
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                startupinfo.wShowWindow = subprocess.SW_HIDE
            
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=timeout,
                encoding='utf-8',
                errors='ignore',
                startupinfo=startupinfo
            )
            result_tuple = (result.returncode == 0, result.stdout, result.stderr)
            
            # 设置缓存
            if use_cache:
                cls._set_cached_result(cmd_tuple, result_tuple)
            
            return result_tuple
        except subprocess.TimeoutExpired:
            return False, "", "Command timeout"
        except FileNotFoundError:
            return False, "", f"ADB not found at: {cmd[0] if cmd else 'unknown'}. Please install Android Platform Tools or set ADB path."
        except Exception as e:
            return False, "", str(e)
    
    @classmethod
    def get_devices(cls) -> List[str]:
        """获取已连接的设备列表 - 使用缓存避免卡顿"""
        success, stdout, stderr = cls.execute_command(['adb', 'devices'], use_cache=True)
        if not success:
            return []
        
        devices = []
        lines = stdout.strip().split('\n')[1:]  # 跳过第一行"List of devices attached"
        for line in lines:
            if line.strip() and '\tdevice' in line:
                device_id = line.split('\t')[0]
                devices.append(device_id)
        return devices
    
    @classmethod
    def get_device_info(cls, serial: str) -> DeviceInfo:
        """获取设备详细信息（支持双卡）"""
        info = DeviceInfo(serial=serial)
        
        # 获取设备型号
        success, stdout, _ = cls.execute_command(
            ['adb', '-s', serial, 'shell', 'getprop', 'ro.product.model']
        )
        if success:
            info.model = stdout.strip()
        
        # 获取设备品牌
        success, stdout, _ = cls.execute_command(
            ['adb', '-s', serial, 'shell', 'getprop', 'ro.product.brand']
        )
        if success:
            brand = stdout.strip()
            info.name = f"{brand} {info.model}"
        
        # 获取Android版本
        success, stdout, _ = cls.execute_command(
            ['adb', '-s', serial, 'shell', 'getprop', 'ro.build.version.release']
        )
        if success and stdout.strip():
            info.android_version = stdout.strip()
        
        # ===== 获取双卡信息 =====
        # 卡一信息
        info.sim1 = cls._get_sim_info(serial, 0)
        # 卡二信息
        info.sim2 = cls._get_sim_info(serial, 1)
        
        # 兼容旧版本：使用卡一作为主卡信息
        info.sim_state = info.sim1.state
        info.sim_operator = info.sim1.operator
        info.phone_number = info.sim1.phone_number
        
        return info
    
    @classmethod
    def _get_sim_info(cls, serial: str, slot: int) -> SimInfo:
        """获取指定SIM卡槽的信息 - 优化版"""
        sim = SimInfo(slot=slot)
        slot_str = str(slot)
        
        # SIM卡状态 - 支持多SIM设备的不同属性名
        state_props = [
            f'gsm.sim.state.slot{slot}',
            f'gsm.sim{slot}.state',
            f'gsm.sim.state',
            f'persist.radio.sim.state',
        ]
        for prop in state_props:
            success, stdout, _ = cls.execute_command(
                ['adb', '-s', serial, 'shell', 'getprop', prop]
            )
            if success and stdout.strip():
                states = stdout.strip().split(',')
                if slot < len(states):
                    state = states[slot].strip()
                else:
                    state = states[0].strip() if states else 'UNKNOWN'
                
                state_map = {
                    'READY': '就绪',
                    'PIN_REQUIRED': '需要PIN码',
                    'PUK_REQUIRED': '需要PUK码',
                    'NETWORK_LOCKED': '网络锁定',
                    'ABSENT': '未插入',
                    'UNKNOWN': '未知',
                    '': '未插入'
                }
                sim.state = state_map.get(state, state)
                break
        
        # 运营商 - 支持多SIM设备
        operator_props = [
            f'gsm.sim.operator.alpha.slot{slot}',
            f'gsm.sim{slot}.operator.alpha',
            f'gsm.sim.operator.alpha',
            f'gsm.operator.alpha',
        ]
        for prop in operator_props:
            success, stdout, _ = cls.execute_command(
                ['adb', '-s', serial, 'shell', 'getprop', prop]
            )
            if success and stdout.strip():
                operators = stdout.strip().split(',')
                if slot < len(operators):
                    sim.operator = operators[slot].strip()
                else:
                    sim.operator = operators[0].strip() if operators else 'Unknown'
                if sim.operator and sim.operator != 'Unknown':
                    break
        
        # 电话号码 - 使用多种方法尝试获取
        phone_number = None
        
        # 方法1: iphonesubinfo service
        service_names = ['iphonesubinfo']
        if slot == 1:
            service_names.append('iphonesubinfo2')
        
        for service in service_names:
            cmd = ['adb', '-s', serial, 'shell', 'service', 'call', service, '13']
            success, stdout, _ = cls.execute_command(cmd, timeout=5)
            if success and stdout:
                # 解析返回的 Parcel 数据
                # 格式类似: Result: Parcel(0x...) 或包含 's16' 和号码
                match = re.search(r"'(\d+)'", stdout)
                if match and len(match.group(1)) > 5:
                    phone_number = match.group(1)
                    break
                # 尝试另一种格式
                lines = stdout.split('\n')
                for line in lines:
                    if re.match(r'^\d{11}$', line.strip()):
                        phone_number = line.strip()
                        break
        
        # 方法2: 使用 telephony 数据库查询（部分设备支持）
        if not phone_number:
            cmd = ['adb', '-s', serial, 'shell', 'content', 'query', 
                   '--uri', f'content://telephony/siminfo/{slot+1}']
            success, stdout, _ = cls.execute_command(cmd, timeout=5)
            if success and stdout:
                match = re.search(r'number=(\d+)', stdout)
                if match:
                    phone_number = match.group(1)
        
        # 方法3: 从 SIM 卡信息中获取
        if not phone_number:
            cmd = ['adb', '-s', serial, 'shell', 'getprop', f'gsm.sim.msisdn.slot{slot}']
            success, stdout, _ = cls.execute_command(cmd)
            if success and stdout.strip():
                phone_number = stdout.strip()
        
        sim.phone_number = phone_number if phone_number else "未知"
        
        # 信号强度 - 改进的信号读取
        signal_level = -1
        
        # 方法1: dumpsys telephony.registry
        success, stdout, _ = cls.execute_command(
            ['adb', '-s', serial, 'shell', 'dumpsys', 'telephony.registry'],
            timeout=5
        )
        if success and stdout:
            # 针对 slot 的匹配
            patterns = [
                rf'CellSignalStrength[^\n]*{{[^}}]*mLevel=(\d+)[^}}]*slot={slot}',
                rf'CellSignalStrength[^\n]*slot={slot}[^\n]*mLevel=(\d+)',
                rf'mSignalStrength[^\n]*slot={slot}[^\n]*level[:=](\d+)',
                rf'\[slot={slot}\][^\n]*CellSignalStrength[^\n]*level[:=](\d+)',
                # 更宽松的匹配
                rf'mSignalStrength.*?level=(\d+)',
                rf'level[:=]\s*(\d+)',
            ]
            for pattern in patterns:
                match = re.search(pattern, stdout, re.DOTALL)
                if match:
                    try:
                        level = int(match.group(1))
                        if 0 <= level <= 4:
                            signal_level = level
                            break
                    except:
                        pass
        
        # 方法2: 通过 settings 获取
        if signal_level < 0:
            cmd = ['adb', '-s', serial, 'shell', 'settings', 'get', 'global', 
                   f'mobile_data_signal_strength_slot_{slot}']
            success, stdout, _ = cls.execute_command(cmd, timeout=3)
            if success and stdout.strip():
                try:
                    signal_level = int(stdout.strip())
                except:
                    pass
        
        sim.signal_level = signal_level if signal_level >= 0 else 0
        
        # 网络类型 - 支持多SIM
        network_type = 'Unknown'
        network_props = [
            f'gsm.network.type.slot{slot}',
            f'gsm.network.type',
            f'gsm.operator.network.type',
        ]
        for prop in network_props:
            success, stdout, _ = cls.execute_command(
                ['adb', '-s', serial, 'shell', 'getprop', prop]
            )
            if success and stdout.strip():
                types = stdout.strip().split(',')
                if slot < len(types):
                    network = types[slot].strip()
                else:
                    network = types[0].strip() if types else 'Unknown'
                
                network_map = {
                    'LTE': '4G',
                    'NR': '5G',
                    'NSA': '5G',
                    'UMTS': '3G',
                    'HSPA': '3G',
                    'HSPAP': '3G',
                    'EDGE': '2G',
                    'GPRS': '2G',
                    'CDMA': '2G',
                }
                network_type = network_map.get(network, network)
                if network_type != 'Unknown':
                    break
        
        sim.network_type = network_type
        return sim
    
    @classmethod
    def make_call(cls, serial: str, phone_number: str, sim_slot: int = 0) -> bool:
        """拨打电话 - 使用service call直接拨号"""
        import time
        
        # 清除可能存在的旧通话状态
        cls.execute_command(['adb', '-s', serial, 'shell', 'input', 'keyevent', 'KEYCODE_ENDCALL'], timeout=2)
        time.sleep(0.5)
        
        # 方法1: 使用 service call phone 直接拨号（最可靠）
        # 2 表示 CALL 操作，s16 表示字符串参数
        try:
            cmd = ['adb', '-s', serial, 'shell', 'service', 'call', 'phone', '2', 's16', phone_number]
            success, stdout, stderr = cls.execute_command(cmd, timeout=5)
            if success and ('Parcel' in stdout or 'Result' in stdout):
                # 等待一下确保拨号已发起
                time.sleep(1)
                return True
        except Exception as e:
            print(f"service call phone 失败: {e}")
        
        # 方法2: 使用 am start 打开拨号界面并模拟点击
        try:
            # 先打开拨号界面
            cmd = [
                'adb', '-s', serial, 'shell', 'am', 'start',
                '-a', 'android.intent.action.DIAL',
                '-d', f'tel:{phone_number}'
            ]
            cls.execute_command(cmd, timeout=5)
            time.sleep(1)
            
            # 模拟点击拨号键（坐标可能因设备而异，尝试几种常见方案）
            # 方案A: 使用 keyevent
            cls.execute_command(['adb', '-s', serial, 'shell', 'input', 'keyevent', 'KEYCODE_CALL'], timeout=2)
            time.sleep(0.5)
            
            return True
        except Exception as e:
            print(f"备选拨号方案失败: {e}")
        
        # 方法3: 使用 input text 和 keyevent（某些设备有效）
        try:
            # 直接通过 keyevent 序列拨号
            for digit in phone_number:
                if digit.isdigit():
                    keycode = f"KEYCODE_{digit}"
                    cls.execute_command(['adb', '-s', serial, 'shell', 'input', 'keyevent', keycode], timeout=1)
                    time.sleep(0.1)
                elif digit == '*':
                    cls.execute_command(['adb', '-s', serial, 'shell', 'input', 'keyevent', 'KEYCODE_STAR'], timeout=1)
                    time.sleep(0.1)
                elif digit == '#':
                    cls.execute_command(['adb', '-s', serial, 'shell', 'input', 'keyevent', 'KEYCODE_POUND'], timeout=1)
                    time.sleep(0.1)
            
            # 按拨号键
            time.sleep(0.5)
            cls.execute_command(['adb', '-s', serial, 'shell', 'input', 'keyevent', 'KEYCODE_CALL'], timeout=2)
            return True
        except Exception as e:
            print(f"KeyEvent拨号失败: {e}")
        
        return False
    
    @classmethod
    def end_call(cls, serial: str) -> bool:
        """挂断电话"""
        # 使用input keyevent模拟挂断键
        cmd = ['adb', '-s', serial, 'shell', 'input', 'keyevent', 'KEYCODE_ENDCALL']
        success, _, _ = cls.execute_command(cmd, timeout=5)
        
        # 如果上述方法失败，尝试另一种方法
        if not success:
            cmd = ['adb', '-s', serial, 'shell', 'input', 'keyevent', '6']
            success, _, _ = cls.execute_command(cmd, timeout=5)
        
        return success
    
    @classmethod
    def get_call_state(cls, serial: str) -> str:
        """获取通话状态"""
        cmd = ['adb', '-s', serial, 'shell', 'dumpsys', 'telephony.registry']
        success, stdout, _ = cls.execute_command(cmd)
        if success:
            if 'mCallState=1' in stdout or 'CALL_STATE_OFFHOOK' in stdout:
                return "通话中"
            elif 'mCallState=2' in stdout or 'CALL_STATE_RINGING' in stdout:
                return "响铃中"
            else:
                return "空闲"
        return "未知"
    
    @classmethod
    def send_sms(cls, serial: str, phone_number: str, message: str, sim_slot: int = 0) -> bool:
        """
        发送短信
        
        Args:
            serial: 设备序列号
            phone_number: 目标手机号
            message: 短信内容
            sim_slot: SIM卡槽 (0=卡一, 1=卡二)
        
        Returns:
            bool: 是否发送成功
        """
        try:
            # 方法1: 使用am start发送短信（打开短信应用预填充）
            # 这种方式需要用户手动点击发送，但最稳定
            escaped_msg = message.replace('"', '\\"').replace("'", "\\'")
            cmd = [
                'adb', '-s', serial, 'shell', 'am', 'start',
                '-a', 'android.intent.action.SENDTO',
                '-d', f'smsto:{phone_number}',
                '--es', 'sms_body', escaped_msg,
                '--ez', 'exit_on_sent', 'true'
            ]
            
            # 尝试指定SIM卡（Android 5.1+ 部分支持）
            if sim_slot > 0:
                cmd.extend(['--ei', 'simId', str(sim_slot)])
            
            success, stdout, stderr = cls.execute_command(cmd, timeout=10)
            
            if success:
                return True
            
            # 方法2: 使用service call isms发送（需要系统权限，通常只适用于系统应用或root）
            # 尝试通过service call直接发送
            return cls._send_sms_via_service(serial, phone_number, message, sim_slot)
            
        except Exception as e:
            print(f"发送短信失败: {e}")
            return False
    
    @classmethod
    def _send_sms_via_service(cls, serial: str, phone_number: str, message: str, sim_slot: int = 0) -> bool:
        """通过service call发送短信（需要更高权限）"""
        try:
            # 将消息转换为UTF-16编码的PDU格式
            # 注意：这种方法在不同设备上可能有兼容性问题
            import binascii
            
            # 简单的7-bit GSM编码转换
            msg_bytes = message.encode('utf-16-be')
            msg_hex = binascii.hexlify(msg_bytes).decode('ascii')
            
            # 构建service call命令
            # isms sendTextForSubscriber [subId] [callingPkg] [destAddr] [scAddr] [text] [sentIntent] [deliveryIntent]
            cmd = [
                'adb', '-s', serial, 'shell', 'service', 'call', 'isms', '7',
                'i32', str(sim_slot),  # subId
                's16', 'com.android.phone',  # callingPkg
                's16', phone_number,  # destAddr
                's16', 'null',  # scAddr
                's16', message,  # text
                's16', 'null',  # sentIntent
                's16', 'null'   # deliveryIntent
            ]
            
            success, stdout, stderr = cls.execute_command(cmd, timeout=10)
            return success and 'Result: Parcel' in stdout
            
        except Exception as e:
            print(f"Service方式发送失败: {e}")
            return False
    
    @classmethod
    def send_sms_direct(cls, serial: str, phone_number: str, message: str, sim_slot: int = 0) -> dict:
        """
        发送短信（增强版，返回详细结果）
        
        Returns:
            dict: 包含success, method, error等信息
        """
        result = {
            'success': False,
            'method': None,
            'error': None,
            'serial': serial,
            'phone_number': phone_number,
            'sim_slot': sim_slot
        }
        
        try:
            # 方法1: 尝试直接通过service发送
            if cls._send_sms_via_service(serial, phone_number, message, sim_slot):
                result['success'] = True
                result['method'] = 'service_call'
                return result
            
            # 方法2: 使用Intent打开短信应用
            escaped_msg = message.replace('"', '\\"').replace("'", "\\'")
            cmd = [
                'adb', '-s', serial, 'shell', 'am', 'start',
                '-a', 'android.intent.action.SENDTO',
                '-d', f'smsto:{phone_number}',
                '--es', 'sms_body', escaped_msg
            ]
            
            success, stdout, stderr = cls.execute_command(cmd, timeout=10)
            
            if success:
                result['success'] = True
                result['method'] = 'intent'
                result['note'] = '已打开短信应用，请手动点击发送'
            else:
                result['error'] = stderr or 'Unknown error'
                
        except Exception as e:
            result['error'] = str(e)
        
        return result


class CallWorker(QThread):
    """电话拨打工作线程"""
    
    log_signal = pyqtSignal(str, str)  # 日志消息，类型(info/success/error/warning)
    progress_signal = pyqtSignal(int, int)  # 当前次数，总次数
    status_signal = pyqtSignal(str)  # 状态更新
    finished_signal = pyqtSignal()
    location_signal = pyqtSignal(str, int, int)  # 位置记录信号：号码，当前次数，总次数
    result_signal = pyqtSignal(dict)  # 结果记录信号：详细拨打记录
    
    def __init__(self, serial: str, phone_number: str, duration: int, count: int, 
                 sim_card: str = "卡一", ping_enabled: bool = False, ping_sim: str = ""):
        super().__init__()
        self.serial = serial
        self.phone_number = phone_number
        self.duration = duration
        self.count = count
        self.sim_card = sim_card  # 使用的SIM卡
        self.ping_enabled = ping_enabled  # 是否启用Ping
        self.ping_sim = ping_sim  # Ping使用的SIM卡
        self.is_running = True
        self.current_count = 0
    
    def run(self):
        """执行拨打任务并记录详细结果"""
        from datetime import datetime
        
        self.log_signal.emit(f"开始拨打测试 - 目标号码: {self.phone_number}, "
                            f"通话时长: {self.duration}秒, 拨打次数: {self.count}", "info")
        
        for i in range(1, self.count + 1):
            if not self.is_running:
                self.log_signal.emit("拨打任务已手动停止", "warning")
                break
            
            self.current_count = i
            self.progress_signal.emit(i, self.count)
            
            # 记录开始时间
            start_time = datetime.now()
            
            self.log_signal.emit(f"\n{'='*50}", "info")
            self.log_signal.emit(f"第 {i}/{self.count} 次拨打", "info")
            self.status_signal.emit(f"正在拨打... ({i}/{self.count})")
            
            # 获取当前信号状态
            signal_info = self._get_signal_info()
            
            # 拨打电话
            self.log_signal.emit(f"正在拨打: {self.phone_number}", "info")
            # 将 sim_card 字符串转换为 slot 数字
            sim_slot = 0 if self.sim_card == "卡一" else 1
            call_success = ADBHelper.make_call(self.serial, self.phone_number, sim_slot)
            
            if not call_success:
                self.log_signal.emit("拨打电话失败！", "error")
                # 发送失败结果
                result_record = {
                    'index': i,
                    'time': start_time.strftime("%H:%M:%S"),
                    'phone_number': self.phone_number,
                    'sim_card': self.sim_card,
                    'ping_status': f"{self.ping_sim}进行中" if self.ping_enabled else "无",
                    'call_result': '失败(拨号失败)',
                    'signal_status': signal_info['text'],
                    'signal_level': signal_info['level'],
                    'duration': 0,
                    'remark': '拨号失败'
                }
                self.result_signal.emit(result_record)
                continue
            
            self.log_signal.emit("电话已拨出", "success")
            
            # 等待电话接通
            self.log_signal.emit("等待接通...", "info")
            connected = False
            wait_time = 0
            max_wait = 30  # 最长等待30秒
            
            while wait_time < max_wait and self.is_running:
                time.sleep(1)
                wait_time += 1
                state = ADBHelper.get_call_state(self.serial)
                if state == "通话中":
                    connected = True
                    self.log_signal.emit("电话已接通", "success")
                    # 发射位置记录信号
                    self.location_signal.emit(self.phone_number, i, self.count)
                    break
            
            if not self.is_running:
                self.log_signal.emit("拨打任务已停止", "warning")
                ADBHelper.end_call(self.serial)
                # 发送停止结果
                result_record = {
                    'index': i,
                    'time': start_time.strftime("%H:%M:%S"),
                    'phone_number': self.phone_number,
                    'sim_card': self.sim_card,
                    'ping_status': f"{self.ping_sim}进行中" if self.ping_enabled else "无",
                    'call_result': '已停止',
                    'signal_status': signal_info['text'],
                    'signal_level': signal_info['level'],
                    'duration': wait_time,
                    'remark': '用户手动停止'
                }
                self.result_signal.emit(result_record)
                break
            
            if not connected:
                self.log_signal.emit("等待接通超时，挂断电话", "warning")
                ADBHelper.end_call(self.serial)
                # 发送超时结果
                result_record = {
                    'index': i,
                    'time': start_time.strftime("%H:%M:%S"),
                    'phone_number': self.phone_number,
                    'sim_card': self.sim_card,
                    'ping_status': f"{self.ping_sim}进行中" if self.ping_enabled else "无",
                    'call_result': '超时(未接通)',
                    'signal_status': signal_info['text'],
                    'signal_level': signal_info['level'],
                    'duration': wait_time,
                    'remark': f'等待{wait_time}秒未接通'
                }
                self.result_signal.emit(result_record)
                time.sleep(2)
                continue
            
            # 保持通话
            self.status_signal.emit(f"通话中... ({i}/{self.count})")
            self.log_signal.emit(f"保持通话 {self.duration} 秒...", "info")
            
            elapsed = 0
            while elapsed < self.duration and self.is_running:
                time.sleep(1)
                elapsed += 1
                if elapsed % 5 == 0:
                    self.log_signal.emit(f"通话中... {elapsed}/{self.duration} 秒", "info")
            
            if not self.is_running:
                self.log_signal.emit("拨打任务已停止，挂断电话", "warning")
                ADBHelper.end_call(self.serial)
                # 发送停止结果
                result_record = {
                    'index': i,
                    'time': start_time.strftime("%H:%M:%S"),
                    'phone_number': self.phone_number,
                    'sim_card': self.sim_card,
                    'ping_status': f"{self.ping_sim}进行中" if self.ping_enabled else "无",
                    'call_result': '已停止',
                    'signal_status': signal_info['text'],
                    'signal_level': signal_info['level'],
                    'duration': elapsed,
                    'remark': '通话中用户停止'
                }
                self.result_signal.emit(result_record)
                break
            
            # 挂断电话
            self.log_signal.emit("挂断电话", "info")
            hangup_success = ADBHelper.end_call(self.serial)
            
            if hangup_success:
                self.log_signal.emit("电话已挂断", "success")
            else:
                self.log_signal.emit("挂断电话可能失败", "warning")
            
            # 发送成功结果
            result_record = {
                'index': i,
                'time': start_time.strftime("%H:%M:%S"),
                'phone_number': self.phone_number,
                'sim_card': self.sim_card,
                'ping_status': f"{self.ping_sim}进行中" if self.ping_enabled else "无",
                'call_result': '成功',
                'signal_status': signal_info['text'],
                'signal_level': signal_info['level'],
                'duration': self.duration,
                'remark': '通话正常完成'
            }
            self.result_signal.emit(result_record)
            
            # 等待2秒再进行下一次拨打
            if i < self.count:
                self.log_signal.emit("等待2秒后进行下一次拨打...", "info")
                time.sleep(2)
        
        self.status_signal.emit("就绪")
        self.log_signal.emit(f"\n{'='*50}", "info")
        self.log_signal.emit("拨打测试完成", "success")
        self.finished_signal.emit()
    
    def _get_signal_info(self) -> dict:
        """获取当前信号信息"""
        try:
            # 获取信号强度
            success, stdout, _ = ADBHelper.execute_command(
                ['adb', '-s', self.serial, 'shell', 'dumpsys', 'telephony.registry'],
                timeout=5
            )
            
            if success and stdout:
                # 解析信号强度
                level = 0
                if 'mSignalStrength' in stdout:
                    try:
                        # 尝试提取信号等级
                        import re
                        match = re.search(r'level[=:]\s*(\d)', stdout)
                        if match:
                            level = int(match.group(1))
                    except:
                        pass
                
                # 根据等级返回描述
                if level >= 4:
                    return {'level': level, 'text': f'强({level}/4)'}
                elif level >= 3:
                    return {'level': level, 'text': f'良好({level}/4)'}
                elif level >= 2:
                    return {'level': level, 'text': f'一般({level}/4)'}
                elif level >= 1:
                    return {'level': level, 'text': f'弱({level}/4)'}
                else:
                    return {'level': level, 'text': '无信号'}
            
            return {'level': 0, 'text': '未知'}
        except:
            return {'level': 0, 'text': '获取失败'}
    
    def stop(self):
        """停止拨打任务"""
        self.is_running = False


class SMSWorker(QThread):
    """短信发送工作线程"""
    
    log_signal = pyqtSignal(str, str)  # 日志消息，类型
    progress_signal = pyqtSignal(int, int)  # 当前次数，总次数
    status_signal = pyqtSignal(str)  # 状态更新
    finished_signal = pyqtSignal()
    result_signal = pyqtSignal(dict)  # 结果记录信号
    
    def __init__(self, serial: str, phone_number: str, message: str, 
                 sim_slot: int = 0, count: int = 1):
        super().__init__()
        self.serial = serial
        self.phone_number = phone_number
        self.message = message
        self.sim_slot = sim_slot
        self.count = count
        self.is_running = True
        self.current_count = 0
    
    def run(self):
        """执行短信发送任务"""
        from datetime import datetime
        
        self.log_signal.emit(f"开始短信发送测试 - 目标号码: {self.phone_number}, "
                            f"内容: {self.message[:30]}..., 次数: {self.count}", "info")
        
        for i in range(1, self.count + 1):
            if not self.is_running:
                self.log_signal.emit("短信发送任务已手动停止", "warning")
                break
            
            self.current_count = i
            self.progress_signal.emit(i, self.count)
            
            start_time = datetime.now()
            
            self.log_signal.emit(f"\n{'='*50}", "info")
            self.log_signal.emit(f"第 {i}/{self.count} 次发送", "info")
            self.status_signal.emit(f"正在发送短信... ({i}/{self.count})")
            
            # 记录位置
            self._record_location()
            
            # 发送短信
            self.log_signal.emit(f"正在发送短信到: {self.phone_number}", "info")
            result = ADBHelper.send_sms_direct(
                self.serial, 
                self.phone_number, 
                self.message,
                self.sim_slot
            )
            
            # 构建结果记录
            result_record = {
                'index': i,
                'time': start_time.strftime("%H:%M:%S"),
                'phone_number': self.phone_number,
                'sim_card': f"卡{self.sim_slot + 1}",
                'test_type': 'sms',
                'sms_content': self.message,
                'ping_status': '无',
                'call_result': '成功' if result['success'] else '失败',
                'signal_status': '发送完成',
                'signal_level': 0,
                'duration': 0,
                'remark': result.get('note', result.get('error', '')) if not result['success'] else '短信已发送'
            }
            
            if result['success']:
                if result.get('method') == 'intent':
                    self.log_signal.emit("已打开短信应用，请手动确认发送", "warning")
                    result_record['remark'] = '已打开短信应用，需手动发送'
                else:
                    self.log_signal.emit("短信发送成功", "success")
                self.result_signal.emit(result_record)
            else:
                self.log_signal.emit(f"短信发送失败: {result.get('error', 'Unknown error')}", "error")
                self.result_signal.emit(result_record)
            
            if not self.is_running:
                break
            
            # 间隔2秒
            if i < self.count:
                self.log_signal.emit("等待2秒后进行下一次发送...", "info")
                time.sleep(2)
        
        self.status_signal.emit("就绪")
        self.log_signal.emit(f"\n{'='*50}", "info")
        self.log_signal.emit("短信发送测试完成", "success")
        self.finished_signal.emit()
    
    def _record_location(self):
        """记录当前位置"""
        # 这里可以通过信号通知主线程记录位置
        pass
    
    def stop(self):
        """停止发送任务"""
        self.is_running = False


class StatusBadge(QLabel):
    """状态标签组件"""
    def __init__(self, text: str = "", status: str = "default", parent=None):
        super().__init__(text, parent)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setStyleSheet(self._get_style(status))
    
    def _get_style(self, status: str) -> str:
        styles = {
            'success': f"""
                QLabel {{
                    background-color: {COLORS['success']}20;
                    color: {COLORS['success']};
                    border: 1px solid {COLORS['success']};
                    border-radius: 12px;
                    padding: 4px 12px;
                    font-weight: bold;
                    font-size: 12px;
                }}
            """,
            'warning': f"""
                QLabel {{
                    background-color: {COLORS['warning']}20;
                    color: {COLORS['warning']};
                    border: 1px solid {COLORS['warning']};
                    border-radius: 12px;
                    padding: 4px 12px;
                    font-weight: bold;
                    font-size: 12px;
                }}
            """,
            'error': f"""
                QLabel {{
                    background-color: {COLORS['error']}20;
                    color: {COLORS['error']};
                    border: 1px solid {COLORS['error']};
                    border-radius: 12px;
                    padding: 4px 12px;
                    font-weight: bold;
                    font-size: 12px;
                }}
            """,
            'info': f"""
                QLabel {{
                    background-color: {COLORS['info']}20;
                    color: {COLORS['info']};
                    border: 1px solid {COLORS['info']};
                    border-radius: 12px;
                    padding: 4px 12px;
                    font-weight: bold;
                    font-size: 12px;
                }}
            """,
            'default': f"""
                QLabel {{
                    background-color: {COLORS['divider']};
                    color: {COLORS['text_secondary']};
                    border: 1px solid {COLORS['border']};
                    border-radius: 12px;
                    padding: 4px 12px;
                    font-weight: bold;
                    font-size: 12px;
                }}
            """
        }
        return styles.get(status, styles['default'])
    
    def set_status(self, status: str):
        self.setStyleSheet(self._get_style(status))


class CardFrame(QFrame):
    """卡片式框架组件"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("cardFrame")
        self.setStyleSheet(f"""
            QFrame#cardFrame {{
                background-color: {COLORS['card_bg']};
                border-radius: 12px;
                border: 1px solid {COLORS['border']};
            }}
        """)
        # 添加阴影效果
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(20)
        shadow.setColor(QColor(0, 0, 0, 30))
        shadow.setOffset(0, 4)
        self.setGraphicsEffect(shadow)


class IconButton(QPushButton):
    """带图标的按钮组件"""
    def __init__(self, icon: str, text: str, color: str = COLORS['primary'], parent=None):
        super().__init__(f"{icon} {text}", parent)
        self.color = color
        self.update_style()
    
    def update_style(self):
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {self.color};
                color: white;
                padding: 12px 24px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 8px;
                border: none;
            }}
            QPushButton:hover {{
                background-color: {self.color}dd;
            }}
            QPushButton:pressed {{
                background-color: {self.color}bb;
            }}
            QPushButton:disabled {{
                background-color: {COLORS['text_secondary']}60;
                color: white;
            }}
        """)


class MainWindow(QMainWindow):
    """主窗口"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{ICONS['phone']} Android电话自动拨打测试工具")
        
        # 设置窗口大小和最小尺寸（确保可以调整大小）
        self.setGeometry(100, 100, 1300, 850)
        self.setMinimumSize(1000, 700)  # 最小尺寸
        
        # 确保窗口可以调整大小
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMinMaxButtonsHint)
        
        self.current_device: Optional[DeviceInfo] = None
        self.call_worker: Optional[CallWorker] = None
        self.device_check_timer: Optional[QTimer] = None
        
        # 日志去重机制：记录重复消息的最后显示时间
        self._last_log_message = {}  # {消息内容: 最后显示时间戳}
        self._log_dedup_interval = 30  # 重复消息最小间隔(秒)
        
        # 关键事件列表 - 这些事件会高亮显示且不受去重限制
        self._critical_events = [
            # 测试生命周期事件
            "开始拨打测试", "拨打测试完成", "拨打任务已停止", "用户手动停止",
            "拨打任务已手动停止", "测试已停止",
            # 通话状态事件
            "电话已接通", "电话已挂断", "电话已拨出", "通话中", 
            "等待接通", "接通超时", "挂断电话",
            # 设备连接事件
            "设备已连接", "设备已断开", "等待设备连接", 
            "检测到", "个设备", "未检测到设备", "请先安装ADB",
            # 结果状态
            "成功", "失败", "错误", "超时", "异常", "已停止", "完成",
            # Ping测试相关
            "Ping测试", "ping", "网络测试",
            # 策略执行
            "开始执行策略", "策略执行完成", "添加策略", "删除策略"
        ]
        
        # 设备连接状态跟踪（用于检测状态变化）
        self._last_device_count = -1  # -1表示初始状态
        self._last_connection_status = None  # None/"connected"/"disconnected"
        
        self.init_ui()
        self.start_device_check()
    
    def init_ui(self):
        """初始化用户界面"""
        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主布局
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        
        # 创建分割器 - 带样式优化的可移动分割线
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(8)  # 加宽手柄，更容易拖动
        splitter.setStyleSheet(f"""
            QSplitter::handle {{
                background-color: {COLORS['border']};
                border-radius: 4px;
                margin: 4px 2px;
            }}
            QSplitter::handle:hover {{
                background-color: {COLORS['primary']};
            }}
            QSplitter::handle:pressed {{
                background-color: {COLORS['primary_dark']};
            }}
        """)
        main_layout.addWidget(splitter)
        
        # 左侧控制面板 - 使用滚动区域，设置合理的尺寸限制
        left_panel = self.create_left_panel()
        left_panel.setMinimumWidth(320)   # 减小最小宽度
        left_panel.setMaximumWidth(550)   # 限制最大宽度但不要过窄
        splitter.addWidget(left_panel)
        
        # 右侧日志面板 - 自适应填充剩余空间
        right_panel = self.create_right_panel()
        right_panel.setMinimumWidth(300)  # 减小最小宽度，允许更灵活调整
        splitter.addWidget(right_panel)
        
        # 设置分割器比例 (左:右 = 1:2)
        splitter.setSizes([420, 880])
        
        # 保存分割器引用，用于后续操作
        self.main_splitter = splitter
        
        # 连接分割器移动信号，实时显示尺寸
        splitter.splitterMoved.connect(self._on_splitter_moved)
        
        # 状态栏
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage(f"{ICONS['info']} 就绪")
        self.status_bar.setStyleSheet(f"""
            QStatusBar {{
                background-color: {COLORS['card_bg']};
                color: {COLORS['text_secondary']};
                padding: 8px 16px;
                border-top: 1px solid {COLORS['border']};
            }}
        """)
        
        # 添加进度条到状态栏
        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximumWidth(200)
        self.progress_bar.setMaximumHeight(16)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setStyleSheet(f"""
            QProgressBar {{
                border: none;
                border-radius: 8px;
                background-color: {COLORS['divider']};
                text-align: center;
                color: {COLORS['text_secondary']};
            }}
            QProgressBar::chunk {{
                background-color: {COLORS['primary']};
                border-radius: 8px;
            }}
        """)
        self.progress_bar.hide()
        self.status_bar.addPermanentWidget(self.progress_bar)
        
        # 添加版权信息
        copyright_label = QLabel("  © 2026 OTC GROUP  ")
        copyright_label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text_secondary']};
                font-size: 11px;
                padding: 0 8px;
            }}
        """)
        self.status_bar.addPermanentWidget(copyright_label)
        
        # 添加联系信息（小字，鼠标悬停显示）
        contact_label = QLabel("  Contact: Tianyuan.liu@samsung.com  ")
        contact_label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text_secondary']};
                font-size: 9px;
                padding: 0 4px;
            }}
        """)
        contact_label.setToolTip("For technical support and usage inquiries")
        self.status_bar.addPermanentWidget(contact_label)
    
    def create_left_panel(self) -> QWidget:
        """创建左侧控制面板 - 使用滚动区域包装以适应不同窗口大小"""
        # 创建滚动区域
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setStyleSheet("""
            QScrollArea {
                background-color: transparent;
                border: none;
            }
            QScrollBar:vertical {
                background-color: #f0f0f0;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background-color: #c0c0c0;
                border-radius: 6px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #a0a0a0;
            }
        """)
        
        # 创建实际的内容面板
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setSpacing(16)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # ===== 设备选择卡片 =====
        device_card = CardFrame()
        device_layout = QVBoxLayout(device_card)
        device_layout.setSpacing(12)
        device_layout.setContentsMargins(20, 20, 20, 20)
        
        # 卡片标题
        title_layout = QHBoxLayout()
        title_icon = QLabel(ICONS['device'])
        title_icon.setStyleSheet("")
        title_layout.addWidget(title_icon)
        title_label = QLabel("设备连接")
        title_label.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        title_label.setStyleSheet(f"color: {COLORS['text_primary']};")
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        
        # 连接状态指示器
        self.connection_status = StatusBadge("未连接", "default")
        title_layout.addWidget(self.connection_status)
        device_layout.addLayout(title_layout)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet(f"background-color: {COLORS['divider']}; max-height: 1px;")
        device_layout.addWidget(line)
        
        # 设备选择
        self.device_combo = QComboBox()
        self.device_combo.setStyleSheet(f"""
            QComboBox {{
                padding: 10px 12px;
                border: 2px solid {COLORS['border']};
                border-radius: 8px;
                background-color: {COLORS['card_bg']};
                font-size: 13px;
            }}
            QComboBox:hover {{
                border-color: {COLORS['primary_light']};
            }}
            QComboBox:focus {{
                border-color: {COLORS['primary']};
            }}
            QComboBox::drop-down {{
                border: none;
                width: 30px;
            }}
            QComboBox QAbstractItemView {{
                border: 1px solid {COLORS['border']};
                border-radius: 8px;
                selection-background-color: {COLORS['primary_light']};
            }}
        """)
        self.device_combo.currentIndexChanged.connect(self.on_device_selected)
        device_layout.addWidget(self.device_combo)
        
        # 刷新按钮和自动检测开关
        refresh_layout = QHBoxLayout()
        refresh_layout.addStretch()
        
        # 自动检测复选框
        from PyQt6.QtWidgets import QCheckBox
        self.auto_refresh_check = QCheckBox("自动检测")
        self.auto_refresh_check.setChecked(True)
        self.auto_refresh_check.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.auto_refresh_check.stateChanged.connect(self._toggle_auto_refresh)
        refresh_layout.addWidget(self.auto_refresh_check)
        refresh_layout.addSpacing(10)
        
        # 手动指定ADB按钮
        self.set_adb_btn = QPushButton("📁 指定ADB")
        self.set_adb_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['warning']};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['warning']}dd;
            }}
        """)
        self.set_adb_btn.setToolTip("如果自动检测失败，请手动指定adb.exe路径")
        self.set_adb_btn.clicked.connect(self._set_adb_path_manual)
        refresh_layout.addWidget(self.set_adb_btn)
        refresh_layout.addSpacing(10)
        
        self.refresh_btn = IconButton(ICONS['refresh'], "刷新设备", COLORS['info'])
        self.refresh_btn.setStyleSheet(self.refresh_btn.styleSheet().replace("12px 24px", "8px 16px"))
        self.refresh_btn.clicked.connect(self.refresh_devices)
        refresh_layout.addWidget(self.refresh_btn)
        device_layout.addLayout(refresh_layout)
        
        layout.addWidget(device_card)
        
        # ===== 手机信息卡片 =====
        phone_info_card = CardFrame()
        phone_info_layout = QVBoxLayout(phone_info_card)
        phone_info_layout.setSpacing(12)
        phone_info_layout.setContentsMargins(20, 20, 20, 20)
        
        # 标题
        phone_title_layout = QHBoxLayout()
        phone_icon = QLabel(ICONS['phone'])
        phone_icon.setStyleSheet("")
        phone_title_layout.addWidget(phone_icon)
        phone_title = QLabel("手机信息")
        phone_title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        phone_title.setStyleSheet(f"color: {COLORS['text_primary']};")
        phone_title_layout.addWidget(phone_title)
        phone_title_layout.addStretch()
        phone_info_layout.addLayout(phone_title_layout)
        
        # 分隔线
        line2 = QFrame()
        line2.setFrameShape(QFrame.Shape.HLine)
        line2.setStyleSheet(f"background-color: {COLORS['divider']}; max-height: 1px;")
        phone_info_layout.addWidget(line2)
        
        # 基本信息网格
        info_grid = QGridLayout()
        info_grid.setSpacing(12)
        
        # 设备名称
        info_grid.addWidget(QLabel(f"{ICONS['device']} 设备名称:"), 0, 0)
        self.device_name_label = QLabel("未连接")
        self.device_name_label.setStyleSheet(f"""
            QLabel {{
                padding: 8px 12px;
                background-color: {COLORS['background']};
                border-radius: 6px;
                color: {COLORS['text_primary']};
                font-size: 13px;
            }}
        """)
        info_grid.addWidget(self.device_name_label, 0, 1)
        
        # 设备型号
        info_grid.addWidget(QLabel(f"{ICONS['phone']} 设备型号:"), 1, 0)
        self.device_model_label = QLabel("未连接")
        self.device_model_label.setStyleSheet(self.device_name_label.styleSheet())
        info_grid.addWidget(self.device_model_label, 1, 1)
        
        # Android 版本
        info_grid.addWidget(QLabel(f"{ICONS['info']} Android版本:"), 2, 0)
        self.android_version_label = QLabel("未连接")
        self.android_version_label.setStyleSheet(self.device_name_label.styleSheet())
        info_grid.addWidget(self.android_version_label, 2, 1)
        
        phone_info_layout.addLayout(info_grid)
        layout.addWidget(phone_info_card)
        
        # ===== 双卡信息卡片 - 横向并排布局 =====
        sim_card = CardFrame()
        sim_layout = QVBoxLayout(sim_card)
        sim_layout.setSpacing(12)
        sim_layout.setContentsMargins(16, 16, 16, 16)
        
        # 标题
        sim_title_layout = QHBoxLayout()
        sim_icon = QLabel(ICONS['sim'])
        sim_icon.setStyleSheet("")
        sim_title_layout.addWidget(sim_icon)
        sim_title = QLabel("双卡状态")
        sim_title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        sim_title.setStyleSheet(f"color: {COLORS['text_primary']};")
        sim_title_layout.addWidget(sim_title)
        sim_title_layout.addStretch()
        sim_layout.addLayout(sim_title_layout)
        
        # 分隔线
        line3 = QFrame()
        line3.setFrameShape(QFrame.Shape.HLine)
        line3.setStyleSheet(f"background-color: {COLORS['divider']}; max-height: 1px;")
        sim_layout.addWidget(line3)
        
        # 双卡横向并排
        sims_layout = QHBoxLayout()
        sims_layout.setSpacing(12)
        
        # 卡一信息
        sim1_frame = self._create_sim_info_frame_compact("卡一", "sim1")
        sims_layout.addWidget(sim1_frame, 1)
        
        # 卡二信息
        sim2_frame = self._create_sim_info_frame_compact("卡二", "sim2")
        sims_layout.addWidget(sim2_frame, 1)
        
        sim_layout.addLayout(sims_layout)
        layout.addWidget(sim_card)
        
        # ===== 实时状态卡片 =====
        status_card = CardFrame()
        status_layout = QVBoxLayout(status_card)
        status_layout.setSpacing(12)
        status_layout.setContentsMargins(20, 20, 20, 20)
        
        # 标题
        status_title_layout = QHBoxLayout()
        status_icon = QLabel(ICONS['time'])
        status_icon.setStyleSheet("")
        status_title_layout.addWidget(status_icon)
        status_title = QLabel("实时状态")
        status_title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        status_title.setStyleSheet(f"color: {COLORS['text_primary']};")
        status_title_layout.addWidget(status_title)
        status_title_layout.addStretch()
        status_layout.addLayout(status_title_layout)
        
        # 分隔线
        line4 = QFrame()
        line4.setFrameShape(QFrame.Shape.HLine)
        line4.setStyleSheet(f"background-color: {COLORS['divider']}; max-height: 1px;")
        status_layout.addWidget(line4)
        
        # 日期时间显示
        datetime_layout = QVBoxLayout()
        datetime_layout.setSpacing(8)
        
        # 日期
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel(f"{ICONS['calendar']} 日期:"))
        self.date_label = QLabel("--")
        self.date_label.setStyleSheet(f"""
            QLabel {{
                padding: 8px 12px;
                background-color: {COLORS['primary_light']}30;
                border-radius: 6px;
                color: {COLORS['primary']};
                font-size: 14px;
                font-weight: bold;
            }}
        """)
        date_layout.addWidget(self.date_label)
        date_layout.addStretch()
        datetime_layout.addLayout(date_layout)
        
        # 时间
        time_layout = QHBoxLayout()
        time_layout.addWidget(QLabel(f"{ICONS['clock']} 时间:"))
        self.time_label = QLabel("--:--:--")
        self.time_label.setStyleSheet(f"""
            QLabel {{
                padding: 8px 12px;
                background-color: {COLORS['primary_light']}30;
                border-radius: 6px;
                color: {COLORS['primary']};
                font-size: 18px;
                font-weight: bold;
                font-family: 'Consolas', monospace;
            }}
        """)
        time_layout.addWidget(self.time_label)
        time_layout.addStretch()
        datetime_layout.addLayout(time_layout)
        
        status_layout.addLayout(datetime_layout)
        
        # 测试状态
        test_status_layout = QHBoxLayout()
        test_status_layout.addWidget(QLabel("测试状态:"))
        self.test_status_badge = StatusBadge("空闲", "default")
        test_status_layout.addWidget(self.test_status_badge)
        test_status_layout.addStretch()
        status_layout.addLayout(test_status_layout)
        
        layout.addWidget(status_card)
        
        # ===== 快速控制卡片 =====
        control_card = CardFrame()
        control_layout = QVBoxLayout(control_card)
        control_layout.setSpacing(12)
        control_layout.setContentsMargins(20, 20, 20, 20)
        
        # 标题
        control_title_layout = QHBoxLayout()
        control_icon = QLabel(ICONS['call'])
        control_icon.setStyleSheet("")
        control_title_layout.addWidget(control_icon)
        control_title = QLabel("快速控制")
        control_title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        control_title.setStyleSheet(f"color: {COLORS['text_primary']};")
        control_title_layout.addWidget(control_title)
        control_title_layout.addStretch()
        control_layout.addLayout(control_title_layout)
        
        # 分隔线
        line5 = QFrame()
        line5.setFrameShape(QFrame.Shape.HLine)
        line5.setStyleSheet(f"background-color: {COLORS['divider']}; max-height: 1px;")
        control_layout.addWidget(line5)
        
        # 开始/停止按钮
        btn_layout = QHBoxLayout()
        
        self.start_btn = IconButton(ICONS['start'], "开始", COLORS['success'])
        self.start_btn.setMinimumHeight(45)
        self.start_btn.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        self.start_btn.clicked.connect(self.start_calling)
        btn_layout.addWidget(self.start_btn)
        
        self.stop_btn = IconButton(ICONS['stop'], "停止", COLORS['error'])
        self.stop_btn.setMinimumHeight(45)
        self.stop_btn.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        self.stop_btn.clicked.connect(self.stop_calling)
        self.stop_btn.setEnabled(False)
        btn_layout.addWidget(self.stop_btn)
        
        control_layout.addLayout(btn_layout)
        
        # 提示文字
        hint = QLabel('💡 详细设置在右侧"测试策略"Tab')
        hint.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 11px;")
        hint.setWordWrap(True)
        hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        control_layout.addWidget(hint)
        
        layout.addWidget(control_card)
        
        # 添加弹性空间，让内容可以压缩
        layout.addStretch()
        
        # 启动时钟更新
        self._start_clock()
        
        # 将面板放入滚动区域
        scroll_area.setWidget(panel)
        return scroll_area
    
    def _create_sim_info_frame_compact(self, sim_name: str, sim_id: str) -> CardFrame:
        """创建紧凑的SIM卡信息展示框架 - 横向并排优化"""
        frame = CardFrame()
        layout = QVBoxLayout(frame)
        layout.setSpacing(8)
        layout.setContentsMargins(12, 12, 12, 12)
        
        # 标题行 - SIM卡名称和状态
        header = QHBoxLayout()
        sim_title = QLabel(f"{ICONS['sim']} {sim_name}")
        sim_title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        sim_title.setStyleSheet(f"color: {COLORS['text_primary']};")
        header.addWidget(sim_title)
        header.addStretch()
        
        status_badge = StatusBadge("未插入", "default")
        status_badge.setStyleSheet(status_badge.styleSheet().replace("padding: 4px 10px", "padding: 2px 8px"))
        setattr(self, f"{sim_id}_status_badge", status_badge)
        header.addWidget(status_badge)
        layout.addLayout(header)
        
        # 信息区域 - 垂直紧凑排列
        info_style = f"""
            QLabel {{
                padding: 4px 8px;
                background-color: {COLORS['background']};
                border-radius: 4px;
                color: {COLORS['text_primary']};
                font-size: 11px;
            }}
        """
        
        # 运营商
        op_layout = QHBoxLayout()
        op_layout.addWidget(QLabel(f"{ICONS['signal']} 运营商"))
        operator_label = QLabel("--")
        operator_label.setStyleSheet(info_style)
        operator_label.setMinimumWidth(80)
        setattr(self, f"{sim_id}_operator_label", operator_label)
        op_layout.addWidget(operator_label, 1)
        layout.addLayout(op_layout)
        
        # 手机号
        num_layout = QHBoxLayout()
        num_layout.addWidget(QLabel(f"{ICONS['call']} 号码"))
        number_label = QLabel("--")
        number_label.setStyleSheet(info_style)
        number_label.setMinimumWidth(80)
        setattr(self, f"{sim_id}_number_label", number_label)
        num_layout.addWidget(number_label, 1)
        layout.addLayout(num_layout)
        
        # 信号强度
        sig_layout = QHBoxLayout()
        sig_layout.addWidget(QLabel(f"{ICONS['wifi']} 信号"))
        
        signal_frame = QFrame()
        signal_layout = QHBoxLayout(signal_frame)
        signal_layout.setSpacing(2)
        signal_layout.setContentsMargins(0, 0, 0, 0)
        
        signal_bars = []
        for i in range(4):
            bar = QFrame()
            bar.setFixedSize(4, 8 + i * 3)
            bar.setStyleSheet(f"background-color: {COLORS['divider']}; border-radius: 1px;")
            signal_layout.addWidget(bar)
            signal_bars.append(bar)
        setattr(self, f"{sim_id}_signal_bars", signal_bars)
        
        signal_value = QLabel("无")
        signal_value.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 10px;")
        setattr(self, f"{sim_id}_signal_value", signal_value)
        signal_layout.addWidget(signal_value)
        signal_layout.addStretch()
        
        sig_layout.addWidget(signal_frame)
        layout.addLayout(sig_layout)
        
        # 网络类型
        net_layout = QHBoxLayout()
        net_layout.addWidget(QLabel(f"{ICONS['network']} 网络"))
        network_label = QLabel("--")
        network_label.setStyleSheet(info_style)
        network_label.setMinimumWidth(50)
        setattr(self, f"{sim_id}_network_label", network_label)
        net_layout.addWidget(network_label, 1)
        layout.addLayout(net_layout)
        
        return frame

    def _start_clock(self):
        """启动时钟更新"""
        from datetime import datetime
        
        def update_clock():
            now = datetime.now()
            # 使用短格式日期
            self.date_label.setText(now.strftime("%m/%d"))
            self.time_label.setText(now.strftime("%H:%M:%S"))
        
        # 创建定时器
        self.clock_timer = QTimer(self)
        self.clock_timer.timeout.connect(update_clock)
        self.clock_timer.start(1000)  # 每秒更新
        update_clock()  # 立即更新一次
    
    def create_right_panel(self) -> QWidget:
        """创建右侧 Tab 面板 - 自适应布局"""
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(16)
        
        # 创建 Tab 控件，设置自适应
        self.tab_widget = QTabWidget()
        self.tab_widget.setDocumentMode(True)
        self.tab_widget.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Expanding
        )
        self.tab_widget.setStyleSheet(f"""
            QTabWidget::pane {{
                border: 1px solid {COLORS['border']};
                border-radius: 12px;
                background-color: {COLORS['card_bg']};
                padding: 4px;
            }}
            QTabBar::tab {{
                background-color: {COLORS['divider']};
                padding: 12px 20px;
                margin-right: 4px;
                border-radius: 8px 8px 0 0;
                font-size: 13px;
                font-weight: 500;
                color: {COLORS['text_secondary']};
            }}
            QTabBar::tab:selected {{
                background-color: {COLORS['card_bg']};
                color: {COLORS['primary']};
                font-weight: bold;
                border-top: 2px solid {COLORS['primary']};
            }}
            QTabBar::tab:hover:!selected {{
                background-color: {COLORS['primary_light']};
                color: {COLORS['text_primary']};
            }}
        """)
        
        # ===== Tab 1: 测试脚本 =====
        self.script_tab = self._create_script_tab()
        self.tab_widget.addTab(self.script_tab, f"{ICONS['script']} 测试脚本")
        
        # ===== Tab 2: 运行日志 =====
        self.log_tab = self._create_log_tab()
        self.tab_widget.addTab(self.log_tab, f"{ICONS['log']} 运行日志")
        
        # ===== Tab 3: 测试结果 =====
        self.result_tab = self._create_result_tab()
        self.tab_widget.addTab(self.result_tab, f"{ICONS['result']} 测试结果")
        
        # ===== Tab 4: 地图打点 =====
        self.map_tab = self._create_map_tab()
        self.tab_widget.addTab(self.map_tab, f"{ICONS['map']} 地图打点")
        
        # ===== Tab 5: ADB 终端 =====
        self.adb_tab = self._create_adb_tab()
        self.tab_widget.addTab(self.adb_tab, "🖥️ ADB终端")
        
        # 默认选中"运行日志"标签
        self.tab_widget.setCurrentIndex(1)
        
        # TabWidget 填充所有可用空间
        layout.addWidget(self.tab_widget, 1)  # stretch factor = 1
        return panel
    
    def _create_script_tab(self) -> QWidget:
        """创建测试脚本 Tab - 双卡策略配置界面"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)
        
        # 标题栏
        header = QHBoxLayout()
        title = QLabel(f"{ICONS['script']} 双卡拨打策略配置")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {COLORS['text_primary']};")
        header.addWidget(title)
        
        header.addStretch()
        
        # 脚本操作按钮
        self.run_script_btn = IconButton(ICONS['run'], "开始测试", COLORS['success'])
        self.run_script_btn.setStyleSheet(self.run_script_btn.styleSheet().replace("12px 24px", "8px 16px"))
        header.addWidget(self.run_script_btn)
        
        self.save_strategy_btn = IconButton(ICONS['save'], "保存策略", COLORS['primary'])
        self.save_strategy_btn.setStyleSheet(self.save_strategy_btn.styleSheet().replace("12px 24px", "8px 16px"))
        header.addWidget(self.save_strategy_btn)
        
        self.load_strategy_btn = IconButton(ICONS['refresh'], "加载策略", COLORS['info'])
        self.load_strategy_btn.setStyleSheet(self.load_strategy_btn.styleSheet().replace("12px 24px", "8px 16px"))
        header.addWidget(self.load_strategy_btn)
        
        layout.addLayout(header)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet(f"background-color: {COLORS['divider']}; max-height: 1px;")
        layout.addWidget(line)
        
        # 目标号码设置区域
        target_group = CardFrame()
        target_layout = QVBoxLayout(target_group)
        target_layout.setSpacing(12)
        target_layout.setContentsMargins(16, 16, 16, 16)
        
        target_title = QLabel(f"{ICONS['target']} 目标手机设置")
        target_title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        target_layout.addWidget(target_title)
        
        target_form = QHBoxLayout()
        
        # 目标卡一
        target_form.addWidget(QLabel("目标卡一:"))
        self.target_sim1_input = QLineEdit()
        self.target_sim1_input.setPlaceholderText("目标手机卡一号")
        self.target_sim1_input.setStyleSheet(f"""
            QLineEdit {{
                padding: 8px 12px;
                border: 2px solid {COLORS['border']};
                border-radius: 6px;
                font-size: 13px;
                min-width: 140px;
            }}
        """)
        target_form.addWidget(self.target_sim1_input)
        
        # 目标卡二
        target_form.addWidget(QLabel("目标卡二:"))
        self.target_sim2_input = QLineEdit()
        self.target_sim2_input.setPlaceholderText("目标手机卡二号")
        self.target_sim2_input.setStyleSheet(self.target_sim1_input.styleSheet())
        target_form.addWidget(self.target_sim2_input)
        
        target_form.addStretch()
        target_layout.addLayout(target_form)
        layout.addWidget(target_group)
        
        # 策略配置区域
        strategy_group = CardFrame()
        strategy_layout = QVBoxLayout(strategy_group)
        strategy_layout.setSpacing(16)
        strategy_layout.setContentsMargins(16, 16, 16, 16)
        
        strategy_title = QLabel(f"{ICONS['settings']} 添加测试策略")
        strategy_title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        strategy_layout.addWidget(strategy_title)
        
        # 测试类型选择
        test_type_layout = QHBoxLayout()
        test_type_layout.addWidget(QLabel("测试类型:"))
        
        self.test_type_combo = QComboBox()
        self.test_type_combo.addItems(["📞 电话拨打", "💬 短信发送"])
        self.test_type_combo.setStyleSheet(f"""
            QComboBox {{
                padding: 8px 12px;
                border: 2px solid {COLORS['primary']};
                border-radius: 6px;
                min-width: 120px;
                font-weight: bold;
            }}
        """)
        self.test_type_combo.currentTextChanged.connect(self._on_test_type_changed)
        test_type_layout.addWidget(self.test_type_combo)
        test_type_layout.addStretch()
        strategy_layout.addLayout(test_type_layout)
        
        # 策略配置行
        config_layout = QHBoxLayout()
        config_layout.setSpacing(12)
        
        # 本机卡选择
        config_layout.addWidget(QLabel("使用:"))
        self.local_sim_combo = QComboBox()
        self.local_sim_combo.addItems(["本机卡一", "本机卡二"])
        self.local_sim_combo.setStyleSheet(f"""
            QComboBox {{
                padding: 8px 12px;
                border: 2px solid {COLORS['border']};
                border-radius: 6px;
                min-width: 100px;
            }}
        """)
        config_layout.addWidget(self.local_sim_combo)
        
        # 箭头
        arrow = QLabel("➜")
        arrow.setStyleSheet("font-size: 20px; color: #999;")
        config_layout.addWidget(arrow)
        
        # 拨打/发送目标选择
        config_layout.addWidget(QLabel("目标:"))
        self.target_sim_combo = QComboBox()
        self.target_sim_combo.addItems(["目标卡一", "目标卡二", "双卡轮流"])
        self.target_sim_combo.setStyleSheet(self.local_sim_combo.styleSheet())
        config_layout.addWidget(self.target_sim_combo)
        
        # 分隔
        config_layout.addSpacing(20)
        
        # 通话时长（电话模式）
        self.duration_label = QLabel("通话时长:")
        config_layout.addWidget(self.duration_label)
        self.call_duration_spin = QSpinBox()
        self.call_duration_spin.setRange(1, 600)
        self.call_duration_spin.setValue(10)
        self.call_duration_spin.setSuffix(" 秒")
        self.call_duration_spin.setStyleSheet(f"""
            QSpinBox {{
                padding: 8px 12px;
                border: 2px solid {COLORS['border']};
                border-radius: 6px;
                min-width: 80px;
            }}
        """)
        config_layout.addWidget(self.call_duration_spin)
        
        # 短信内容输入（短信模式，默认隐藏）
        self.sms_content_label = QLabel("短信内容:")
        self.sms_content_label.setVisible(False)
        config_layout.addWidget(self.sms_content_label)
        
        self.sms_content_input = QLineEdit()
        self.sms_content_input.setPlaceholderText("请输入短信内容")
        self.sms_content_input.setText("测试短信")
        self.sms_content_input.setVisible(False)
        self.sms_content_input.setStyleSheet(f"""
            QLineEdit {{
                padding: 8px 12px;
                border: 2px solid {COLORS['border']};
                border-radius: 6px;
                min-width: 200px;
            }}
        """)
        config_layout.addWidget(self.sms_content_input)
        
        # 发送次数
        config_layout.addWidget(QLabel("次数:"))
        self.call_count_spin = QSpinBox()
        self.call_count_spin.setRange(1, 100)
        self.call_count_spin.setValue(3)
        self.call_count_spin.setSuffix(" 次")
        self.call_count_spin.setStyleSheet(self.call_duration_spin.styleSheet())
        config_layout.addWidget(self.call_count_spin)
        
        config_layout.addStretch()
        strategy_layout.addLayout(config_layout)
        
        # Ping 包选项
        ping_layout = QHBoxLayout()
        self.ping_check = QLabel("☐")
        self.ping_check.setStyleSheet("font-size: 18px; color: #999; cursor: pointer;")
        self.ping_check.mousePressEvent = self._toggle_ping
        self.ping_enabled = False
        ping_layout.addWidget(self.ping_check)
        
        self.ping_sim_combo = QComboBox()
        self.ping_sim_combo.addItems(["使用本机卡一进行Ping", "使用本机卡二进行Ping"])
        self.ping_sim_combo.setEnabled(False)
        self.ping_sim_combo.setStyleSheet(f"""
            QComboBox {{
                padding: 8px 12px;
                border: 2px solid {COLORS['border']};
                border-radius: 6px;
                min-width: 180px;
            }}
            QComboBox:disabled {{
                background-color: {COLORS['divider']};
                color: {COLORS['text_secondary']};
            }}
        """)
        ping_layout.addWidget(self.ping_sim_combo)
        
        self.ping_target_input = QLineEdit()
        self.ping_target_input.setPlaceholderText("Ping 目标 IP 或域名")
        self.ping_target_input.setText("8.8.8.8")
        self.ping_target_input.setEnabled(False)
        self.ping_target_input.setStyleSheet(f"""
            QLineEdit {{
                padding: 8px 12px;
                border: 2px solid {COLORS['border']};
                border-radius: 6px;
                min-width: 150px;
            }}
            QLineEdit:disabled {{
                background-color: {COLORS['divider']};
            }}
        """)
        ping_layout.addWidget(self.ping_target_input)
        
        ping_layout.addStretch()
        strategy_layout.addLayout(ping_layout)
        
        # 添加策略按钮
        add_btn_layout = QHBoxLayout()
        add_btn_layout.addStretch()
        
        self.add_strategy_btn = IconButton(ICONS['success'], "添加策略", COLORS['success'])
        self.add_strategy_btn.setStyleSheet(self.add_strategy_btn.styleSheet().replace("12px 24px", "8px 24px"))
        self.add_strategy_btn.clicked.connect(self._add_strategy)
        add_btn_layout.addWidget(self.add_strategy_btn)
        
        strategy_layout.addLayout(add_btn_layout)
        layout.addWidget(strategy_group)
        
        # 策略列表区域
        list_title = QLabel(f"{ICONS['list']} 策略列表 (将按顺序执行)")
        list_title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        layout.addWidget(list_title)
        
        # 策略列表容器
        self.strategy_list = QWidget()
        self.strategy_list_layout = QVBoxLayout(self.strategy_list)
        self.strategy_list_layout.setSpacing(8)
        self.strategy_list_layout.setContentsMargins(0, 0, 0, 0)
        self.strategy_list_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        
        # 滚动区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setWidget(self.strategy_list)
        scroll.setStyleSheet(f"""
            QScrollArea {{
                background-color: {COLORS['background']};
                border-radius: 8px;
                border: 1px solid {COLORS['border']};
            }}
        """)
        # 策略列表填充剩余空间
        layout.addWidget(scroll, 1)  # stretch factor = 1
        
        # 底部统计和操作
        bottom_layout = QHBoxLayout()
        
        self.strategy_stats = QLabel("暂无策略")
        self.strategy_stats.setStyleSheet(f"color: {COLORS['text_secondary']};")
        bottom_layout.addWidget(self.strategy_stats)
        
        bottom_layout.addStretch()
        
        self.clear_strategy_btn = IconButton(ICONS['clear'], "清空全部", COLORS['warning'])
        self.clear_strategy_btn.setStyleSheet(self.clear_strategy_btn.styleSheet().replace("12px 24px", "6px 16px"))
        self.clear_strategy_btn.clicked.connect(self._clear_strategies)
        bottom_layout.addWidget(self.clear_strategy_btn)
        
        layout.addLayout(bottom_layout)
        
        # 存储策略列表
        self.strategies = []
        
        return tab
    
    def _on_test_type_changed(self, test_type: str):
        """测试类型改变时的处理"""
        is_sms = "短信" in test_type
        
        # 显示/隐藏相关控件
        self.duration_label.setVisible(not is_sms)
        self.call_duration_spin.setVisible(not is_sms)
        
        self.sms_content_label.setVisible(is_sms)
        self.sms_content_input.setVisible(is_sms)
        
        # 更新Ping选项的可见性（短信模式下不需要Ping）
        if hasattr(self, 'ping_check'):
            self.ping_check.setVisible(not is_sms)
            self.ping_sim_combo.setVisible(not is_sms)
            self.ping_target_input.setVisible(not is_sms)
    
    def _toggle_ping(self, event=None):
        """切换 Ping 选项"""
        self.ping_enabled = not self.ping_enabled
        if self.ping_enabled:
            self.ping_check.setText("☑")
            self.ping_check.setStyleSheet("font-size: 18px; color: #4CAF50;")
            self.ping_sim_combo.setEnabled(True)
            self.ping_target_input.setEnabled(True)
        else:
            self.ping_check.setText("☐")
            self.ping_check.setStyleSheet("font-size: 18px; color: #999;")
            self.ping_sim_combo.setEnabled(False)
            self.ping_target_input.setEnabled(False)
    
    def _add_strategy(self):
        """添加策略到列表"""
        test_type = "sms" if "短信" in self.test_type_combo.currentText() else "call"
        local_sim = self.local_sim_combo.currentText()
        target = self.target_sim_combo.currentText()
        count = self.call_count_spin.value()
        
        # 根据测试类型获取不同参数
        if test_type == "call":
            duration = self.call_duration_spin.value()
            sms_content = None
        else:
            duration = 0  # 短信不需要时长
            sms_content = self.sms_content_input.text().strip()
            if not sms_content:
                QMessageBox.warning(self, f"{ICONS['warning']} 警告", "请输入短信内容！")
                return
        
        # 获取目标号码
        target_num = ""
        if target == "目标卡一":
            target_num = self.target_sim1_input.text() or "未设置"
        elif target == "目标卡二":
            target_num = self.target_sim2_input.text() or "未设置"
        else:  # 双卡轮流
            target_num = f"卡一({self.target_sim1_input.text() or '未设置'}) / 卡二({self.target_sim2_input.text() or '未设置'})"
        
        # 构建策略描述
        ping_info = ""
        if test_type == "call" and self.ping_enabled:
            ping_sim = self.ping_sim_combo.currentText().replace("使用", "").replace("进行Ping", "")
            ping_target = self.ping_target_input.text()
            ping_info = f" + {ping_sim}Ping[{ping_target}]"
        
        strategy = {
            'test_type': test_type,
            'local_sim': local_sim,
            'target': target,
            'target_num': target_num,
            'duration': duration,
            'count': count,
            'sms_content': sms_content,
            'ping_enabled': self.ping_enabled if test_type == "call" else False,
            'ping_sim': self.ping_sim_combo.currentText() if self.ping_enabled and test_type == "call" else "",
            'ping_target': self.ping_target_input.text() if self.ping_enabled and test_type == "call" else ""
        }
        
        self.strategies.append(strategy)
        self._refresh_strategy_list()
        self._update_strategy_stats()
    
    def _refresh_strategy_list(self):
        """刷新策略列表显示"""
        # 清除现有项
        while self.strategy_list_layout.count():
            item = self.strategy_list_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        # 添加策略卡片
        for i, strategy in enumerate(self.strategies):
            card = self._create_strategy_card(i, strategy)
            self.strategy_list_layout.addWidget(card)
        
        self.strategy_list_layout.addStretch()
    
    def _create_strategy_card(self, index: int, strategy: dict) -> CardFrame:
        """创建策略卡片"""
        card = CardFrame()
        layout = QHBoxLayout(card)
        layout.setSpacing(12)
        layout.setContentsMargins(12, 12, 12, 12)
        
        # 序号
        num_label = QLabel(f"#{index + 1}")
        num_label.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        num_label.setStyleSheet(f"color: {COLORS['primary']}; min-width: 30px;")
        layout.addWidget(num_label)
        
        # 测试类型图标
        is_sms = strategy.get('test_type') == 'sms'
        test_type_icon = "💬" if is_sms else "📞"
        test_type_color = COLORS['success'] if is_sms else COLORS['primary']
        
        sim_icon = "📱" if "卡一" in strategy['local_sim'] else "📲"
        target_icon = "🎯"
        
        desc_text = f"{test_type_icon} {sim_icon} {strategy['local_sim']} ➜ {target_icon} {strategy['target']}"
        desc = QLabel(desc_text)
        desc.setFont(QFont("Segoe UI", 12))
        desc.setStyleSheet(f"color: {COLORS['text_primary']}; min-width: 200px;")
        layout.addWidget(desc)
        
        # 目标号码
        num_label = QLabel(f"号码: {strategy['target_num']}")
        num_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 12px;")
        layout.addWidget(num_label)
        
        # 参数（根据测试类型显示不同信息）
        if is_sms:
            sms_preview = strategy.get('sms_content', '')[:20] + "..." if len(strategy.get('sms_content', '')) > 20 else strategy.get('sms_content', '')
            params = QLabel(f"短信: {sms_preview} × {strategy['count']}次")
            params.setStyleSheet(f"color: {COLORS['success']}; font-size: 12px;")
        else:
            params = QLabel(f"通话{strategy['duration']}秒 × {strategy['count']}次")
            params.setStyleSheet(f"color: {COLORS['info']}; font-size: 12px;")
        layout.addWidget(params)
        
        # Ping 标识（仅电话模式）
        if not is_sms and strategy['ping_enabled']:
            ping_badge = StatusBadge("Ping", "info")
            layout.addWidget(ping_badge)
        
        # 短信标识
        if is_sms:
            sms_badge = StatusBadge("SMS", "success")
            layout.addWidget(sms_badge)
        
        layout.addStretch()
        
        # 删除按钮
        delete_btn = QPushButton("🗑️")
        delete_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                border: none;
                font-size: 16px;
                padding: 4px 8px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['error']}20;
                border-radius: 4px;
            }}
        """)
        delete_btn.clicked.connect(lambda: self._remove_strategy(index))
        layout.addWidget(delete_btn)
        
        return card
    
    def _remove_strategy(self, index: int):
        """删除指定策略"""
        if 0 <= index < len(self.strategies):
            self.strategies.pop(index)
            self._refresh_strategy_list()
            self._update_strategy_stats()
    
    def _clear_strategies(self):
        """清空所有策略"""
        self.strategies.clear()
        self._refresh_strategy_list()
        self._update_strategy_stats()
    
    def _update_strategy_stats(self):
        """更新策略统计"""
        if not self.strategies:
            self.strategy_stats.setText("暂无策略")
        else:
            total_count = sum(s['count'] for s in self.strategies)
            call_count = sum(1 for s in self.strategies if s.get('test_type') == 'call')
            sms_count = sum(1 for s in self.strategies if s.get('test_type') == 'sms')
            
            stats_text = f"共 {len(self.strategies)} 个策略"
            if call_count > 0:
                stats_text += f"，电话:{call_count}"
            if sms_count > 0:
                stats_text += f"，短信:{sms_count}"
            stats_text += f"，总计{total_count}次"
            
            self.strategy_stats.setText(stats_text)
    
    def _create_log_tab(self) -> QWidget:
        """创建运行日志 Tab"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)
        
        # 标题栏（带统计信息）
        title_bar = QHBoxLayout()
        
        title = QLabel(f"{ICONS['log']} 运行日志")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {COLORS['text_primary']};")
        title_bar.addWidget(title)
        
        title_bar.addStretch()
        
        # 重置布局按钮
        reset_layout_btn = QPushButton(f"{ICONS['panel']} 重置布局")
        reset_layout_btn.setToolTip("双击分割线也可重置为默认比例")
        reset_layout_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['background']};
                color: {COLORS['text_secondary']};
                border: 1px solid {COLORS['border']};
                border-radius: 6px;
                padding: 4px 10px;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['primary_light']};
                color: {COLORS['primary']};
                border-color: {COLORS['primary']};
            }}
        """)
        reset_layout_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        reset_layout_btn.clicked.connect(self.reset_splitter)
        title_bar.addWidget(reset_layout_btn)
        
        title_bar.addSpacing(12)
        
        # 统计信息标签
        self.stats_label = QLabel(f"{ICONS['info']} 等待开始...")
        self.stats_label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text_secondary']};
                font-size: 13px;
                padding: 4px 12px;
                background-color: {COLORS['background']};
                border-radius: 6px;
            }}
        """)
        title_bar.addWidget(self.stats_label)
        
        layout.addLayout(title_bar)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet(f"background-color: {COLORS['divider']}; max-height: 1px;")
        layout.addWidget(line)
        
        # 日志文本框（保持原有的样式）
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 11))
        self.log_text.setStyleSheet(f"""
            QTextEdit {{
                background-color: #1e1e2e;
                color: #cdd6f4;
                border: none;
                border-radius: 8px;
                padding: 12px;
                selection-background-color: {COLORS['primary']}40;
            }}
            QScrollBar:vertical {{
                background-color: #313244;
                width: 12px;
                border-radius: 6px;
            }}
            QScrollBar::handle:vertical {{
                background-color: #585b70;
                border-radius: 6px;
                min-height: 20px;
            }}
            QScrollBar::handle:vertical:hover {{
                background-color: #6c7086;
            }}
        """)
        # 日志文本框填充剩余空间
        layout.addWidget(self.log_text, 1)  # stretch factor = 1
        
        # 日志操作按钮
        btn_layout = QHBoxLayout()
        
        self.clear_log_btn = IconButton(ICONS['clear'], "清空日志", COLORS['warning'])
        self.clear_log_btn.setStyleSheet(self.clear_log_btn.styleSheet().replace("12px 24px", "8px 16px"))
        self.clear_log_btn.clicked.connect(self.clear_log)
        btn_layout.addWidget(self.clear_log_btn)
        
        self.save_log_btn = IconButton(ICONS['save'], "保存日志", COLORS['primary'])
        self.save_log_btn.setStyleSheet(self.save_log_btn.styleSheet().replace("12px 24px", "8px 16px"))
        self.save_log_btn.clicked.connect(self.save_log)
        btn_layout.addWidget(self.save_log_btn)
        
        self.auto_scroll_check = QLabel("✓ 自动滚动")
        self.auto_scroll_check.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['success']};
                font-size: 12px;
                padding: 4px 8px;
            }}
        """)
        btn_layout.addWidget(self.auto_scroll_check)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        return tab
    
    def _create_result_tab(self) -> QWidget:
        """创建测试结果 Tab - 包含详细拨打记录表格"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)
        
        # 标题栏
        header = QHBoxLayout()
        title = QLabel(f"{ICONS['result']} 测试结果统计")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {COLORS['text_primary']};")
        header.addWidget(title)
        header.addStretch()
        
        self.export_result_btn = IconButton(ICONS['save'], "导出Excel", COLORS['primary'])
        self.export_result_btn.setStyleSheet(self.export_result_btn.styleSheet().replace("12px 24px", "8px 16px"))
        self.export_result_btn.clicked.connect(self._export_results_to_excel)
        header.addWidget(self.export_result_btn)
        
        layout.addLayout(header)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet(f"background-color: {COLORS['divider']}; max-height: 1px;")
        layout.addWidget(line)
        
        # 统计卡片区域
        stats_grid = QHBoxLayout()
        stats_grid.setSpacing(12)
        
        # 总拨打次数
        self.total_calls_card = self._create_stat_card(ICONS['count'], "总拨打次数", "0", COLORS['primary'])
        stats_grid.addWidget(self.total_calls_card)
        
        # 成功次数
        self.success_calls_card = self._create_stat_card(ICONS['success'], "成功次数", "0", COLORS['success'])
        stats_grid.addWidget(self.success_calls_card)
        
        # 失败次数
        self.failed_calls_card = self._create_stat_card(ICONS['error'], "失败次数", "0", COLORS['error'])
        stats_grid.addWidget(self.failed_calls_card)
        
        # 成功率
        self.success_rate_card = self._create_stat_card(ICONS['chart'], "成功率", "0%", COLORS['info'])
        stats_grid.addWidget(self.success_rate_card)
        
        layout.addLayout(stats_grid)
        
        # 详细结果表格
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(7)
        self.result_table.setHorizontalHeaderLabels([
            "序号", "时间", "拨打号码", "使用SIM卡", "Ping状态", "通话结果", "信号状态"
        ])
        self.result_table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {COLORS['card_bg']};
                border: 1px solid {COLORS['border']};
                border-radius: 8px;
                gridline-color: {COLORS['divider']};
            }}
            QTableWidget::item {{
                padding: 8px;
                border-bottom: 1px solid {COLORS['divider']};
            }}
            QHeaderView::section {{
                background-color: {COLORS['primary_light']};
                padding: 10px;
                font-weight: bold;
                border: none;
                border-bottom: 2px solid {COLORS['primary']};
            }}
        """)
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.result_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        # 设置列宽
        self.result_table.setColumnWidth(0, 50)   # 序号
        self.result_table.setColumnWidth(1, 80)   # 时间
        self.result_table.setColumnWidth(2, 120)  # 拨打号码
        self.result_table.setColumnWidth(3, 90)   # SIM卡
        self.result_table.setColumnWidth(4, 100)  # Ping状态
        self.result_table.setColumnWidth(5, 100)  # 通话结果
        self.result_table.setColumnWidth(6, 120)  # 信号状态
        
        # 表格填充剩余空间
        layout.addWidget(self.result_table, 1)  # stretch factor = 1
        
        # 底部操作按钮
        btn_layout = QHBoxLayout()
        
        self.clear_result_btn = IconButton(ICONS['clear'], "清空结果", COLORS['warning'])
        self.clear_result_btn.setStyleSheet(self.clear_result_btn.styleSheet().replace("12px 24px", "8px 16px"))
        self.clear_result_btn.clicked.connect(self._clear_results)
        btn_layout.addWidget(self.clear_result_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # 初始化结果列表
        self.call_results = []
        
        return tab
    
    def _add_result_record(self, record: dict):
        """添加一条测试结果记录到表格"""
        row = self.result_table.rowCount()
        self.result_table.insertRow(row)
        
        # 序号
        self.result_table.setItem(row, 0, QTableWidgetItem(str(record['index'])))
        
        # 时间
        self.result_table.setItem(row, 1, QTableWidgetItem(record['time']))
        
        # 拨打号码
        self.result_table.setItem(row, 2, QTableWidgetItem(record['phone_number']))
        
        # SIM卡
        sim_item = QTableWidgetItem(record['sim_card'])
        if '卡一' in record['sim_card']:
            sim_item.setForeground(QBrush(QColor(COLORS['primary'])))
        else:
            sim_item.setForeground(QBrush(QColor(COLORS['info'])))
        self.result_table.setItem(row, 3, sim_item)
        
        # Ping状态
        ping_item = QTableWidgetItem(record['ping_status'])
        if '进行中' in record['ping_status']:
            ping_item.setBackground(QBrush(QColor(COLORS['primary_light'])))
        elif '无' in record['ping_status']:
            ping_item.setForeground(QBrush(QColor(COLORS['text_secondary'])))
        self.result_table.setItem(row, 4, ping_item)
        
        # 通话结果
        result_item = QTableWidgetItem(record['call_result'])
        if '成功' in record['call_result']:
            result_item.setForeground(QBrush(QColor(COLORS['success'])))
            result_item.setIcon(QIcon())  # 可以设置成功图标
        elif '失败' in record['call_result']:
            result_item.setForeground(QBrush(QColor(COLORS['error'])))
        elif '超时' in record['call_result']:
            result_item.setForeground(QBrush(QColor(COLORS['warning'])))
        self.result_table.setItem(row, 5, result_item)
        
        # 信号状态
        signal_item = QTableWidgetItem(record['signal_status'])
        signal_level = record.get('signal_level', 0)
        if signal_level >= 3:
            signal_item.setForeground(QBrush(QColor(COLORS['success'])))
        elif signal_level >= 2:
            signal_item.setForeground(QBrush(QColor(COLORS['info'])))
        elif signal_level >= 1:
            signal_item.setForeground(QBrush(QColor(COLORS['warning'])))
        self.result_table.setItem(row, 6, signal_item)
        
        # 保存到列表
        self.call_results.append(record)
        
        # 滚动到最新行
        self.result_table.scrollToBottom()
        
        # 更新统计卡片
        self._update_result_stats()
    
    def _update_result_stats(self):
        """更新统计卡片"""
        total = len(self.call_results)
        success = sum(1 for r in self.call_results if '成功' in r['call_result'])
        failed = sum(1 for r in self.call_results if '失败' in r['call_result'] or '超时' in r['call_result'])
        
        self.total_calls_card.value_label.setText(str(total))
        self.success_calls_card.value_label.setText(str(success))
        self.failed_calls_card.value_label.setText(str(failed))
        
        if total > 0:
            rate = (success / total) * 100
            self.success_rate_card.value_label.setText(f"{rate:.1f}%")
        else:
            self.success_rate_card.value_label.setText("0%")
    
    def _clear_results(self):
        """清空结果表格"""
        self.result_table.setRowCount(0)
        self.call_results.clear()
        self.total_calls_card.value_label.setText("0")
        self.success_calls_card.value_label.setText("0")
        self.failed_calls_card.value_label.setText("0")
        self.success_rate_card.value_label.setText("0%")
    
    def _export_results_to_excel(self):
        """导出结果到Excel文件"""
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        from datetime import datetime
        
        if not self.call_results:
            QMessageBox.information(self, "提示", "没有测试结果可导出")
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "导出测试结果",
            f"test_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*.*)"
        )
        
        if not filename:
            return
        
        try:
            if filename.endswith('.csv'):
                # 导出CSV格式
                with open(filename, 'w', encoding='utf-8-sig') as f:
                    # 写入表头
                    f.write("序号,时间,拨打号码,使用SIM卡,Ping状态,通话结果,信号状态,通话时长(秒),备注\n")
                    # 写入数据
                    for r in self.call_results:
                        f.write(f"{r['index']},{r['time']},{r['phone_number']},"
                               f"{r['sim_card']},{r['ping_status']},{r['call_result']},"
                               f"{r['signal_status']},{r.get('duration', 10)},{r.get('remark', '')}\n")
            else:
                # 导出Excel格式，尝试使用openpyxl
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "拨打测试结果"
                    
                    # 设置表头
                    headers = ["序号", "时间", "拨打号码", "使用SIM卡", "Ping状态", "通话结果", "信号状态", "通话时长(秒)", "备注"]
                    ws.append(headers)
                    
                    # 设置表头样式
                    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # 写入数据
                    for r in self.call_results:
                        row = [
                            r['index'],
                            r['time'],
                            r['phone_number'],
                            r['sim_card'],
                            r['ping_status'],
                            r['call_result'],
                            r['signal_status'],
                            r.get('duration', 10),
                            r.get('remark', '')
                        ]
                        ws.append(row)
                    
                    # 设置列宽
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 12
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 12
                    ws.column_dimensions['E'].width = 15
                    ws.column_dimensions['F'].width = 12
                    ws.column_dimensions['G'].width = 15
                    ws.column_dimensions['H'].width = 15
                    ws.column_dimensions['I'].width = 20
                    
                    # 保存
                    wb.save(filename)
                    
                except ImportError:
                    # 如果没有openpyxl，导出为CSV并提示用户
                    csv_filename = filename.replace('.xlsx', '.csv')
                    with open(csv_filename, 'w', encoding='utf-8-sig') as f:
                        f.write("序号,时间,拨打号码,使用SIM卡,Ping状态,通话结果,信号状态,通话时长(秒),备注\n")
                        for r in self.call_results:
                            f.write(f"{r['index']},{r['time']},{r['phone_number']},"
                                   f"{r['sim_card']},{r['ping_status']},{r['call_result']},"
                                   f"{r['signal_status']},{r.get('duration', 10)},{r.get('remark', '')}\n")
                    
                    QMessageBox.information(
                        self, 
                        "导出成功", 
                        f"已导出为CSV格式（未安装openpyxl库）：\n{csv_filename}\n\n"
                        f"如需导出Excel格式，请安装：pip install openpyxl"
                    )
                    return
            
            QMessageBox.information(self, "导出成功", f"测试结果已导出到：\n{filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出时发生错误：\n{str(e)}")
    
    def _create_map_tab(self) -> QWidget:
        """创建地图打点 Tab - 使用 Folium + WebEngine 实现真实地图"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)
        
        # 标题栏
        header = QHBoxLayout()
        title = QLabel(f"{ICONS['map']} 通话位置地图")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {COLORS['text_primary']};")
        header.addWidget(title)
        header.addStretch()
        
        # 通话计数
        self.location_count_label = QLabel("通话记录: 0")
        self.location_count_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 12px;")
        header.addWidget(self.location_count_label)
        
        # 位置状态
        self.location_status = StatusBadge("未定位", "default")
        header.addWidget(self.location_status)
        layout.addLayout(header)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet(f"background-color: {COLORS['divider']}; max-height: 1px;")
        layout.addWidget(line)
        
        # 真实地图 - 使用 QWebEngineView 加载 Folium 地图
        self.map_view = QWebEngineView()
        self.map_view.setMinimumHeight(300)
        self.map_view.setStyleSheet(f"""
            QWebEngineView {{
                border: 2px solid {COLORS['border']};
                border-radius: 12px;
            }}
        """)
        # 初始化空白地图
        self._init_empty_map()
        layout.addWidget(self.map_view, 1)  # stretch factor = 1
        
        # 当前位置信息
        self.current_location_display = QLabel(f"{ICONS['location']} 当前位置: 等待记录...")
        self.current_location_display.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text_secondary']};
                font-size: 13px;
                padding: 8px;
                background-color: {COLORS['background']};
                border-radius: 6px;
            }}
        """)
        layout.addWidget(self.current_location_display)
        
        # 位置列表（文本形式补充）
        self.location_list = QTextEdit()
        self.location_list.setReadOnly(True)
        self.location_list.setMaximumHeight(120)
        self.location_list.setPlaceholderText("通话位置记录将显示在这里...")
        self.location_list.setStyleSheet(f"""
            QTextEdit {{
                background-color: #f8f9fa;
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
                border-radius: 8px;
                padding: 10px;
                font-size: 11px;
            }}
        """)
        layout.addWidget(self.location_list)
        
        # 底部按钮
        btn_layout = QHBoxLayout()
        
        self.refresh_location_btn = IconButton(ICONS['refresh'], "刷新位置", COLORS['info'])
        self.refresh_location_btn.setStyleSheet(self.refresh_location_btn.styleSheet().replace("12px 24px", "8px 16px"))
        self.refresh_location_btn.clicked.connect(self._refresh_current_location)
        btn_layout.addWidget(self.refresh_location_btn)
        
        self.clear_map_btn = IconButton(ICONS['clear'], "清空标记", COLORS['warning'])
        self.clear_map_btn.setStyleSheet(self.clear_map_btn.styleSheet().replace("12px 24px", "8px 16px"))
        self.clear_map_btn.clicked.connect(self._clear_location_records)
        btn_layout.addWidget(self.clear_map_btn)
        
        self.export_map_btn = IconButton(ICONS['save'], "导出轨迹", COLORS['primary'])
        self.export_map_btn.setStyleSheet(self.export_map_btn.styleSheet().replace("12px 24px", "8px 16px"))
        self.export_map_btn.clicked.connect(self._export_location_tracks)
        btn_layout.addWidget(self.export_map_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # 初始化位置记录列表
        self.location_records = []
        
        return tab
    
    def _refresh_current_location(self):
        """刷新当前位置"""
        if not self.current_device:
            self.current_location_display.setText(f"{ICONS['location']} 请先连接设备")
            return
        self.current_location_display.setText(f"{ICONS['location']} 正在获取位置...")
        import threading
        thread = threading.Thread(target=self._get_location_thread)
        thread.daemon = True
        thread.start()
    
    def _get_location_thread(self):
        """在线程中获取位置"""
        location = self._get_device_location()
        self.current_location_display.setText(f"{ICONS['location']} {location}")
    
    def _get_device_location(self) -> str:
        """获取设备位置信息 - 优先获取GPS坐标"""
        if not self.current_device:
            return "未连接设备"
        try:
            # 方法1: 尝试获取GPS位置（最准确）
            success, stdout, _ = ADBHelper.execute_command(
                ['adb', '-s', self.current_device.serial, 'shell', 'dumpsys', 'location'],
                timeout=5
            )
            if success and stdout:
                lines = stdout.strip().split('\n')
                for line in lines:
                    line_lower = line.lower()
                    # 查找包含经纬度的行
                    if ('latitude' in line_lower and 'longitude' in line_lower) or \
                       ('lat=' in line_lower and 'lon=' in line_lower):
                        # 尝试提取坐标
                        import re
                        lat_match = re.search(r'lat(?:itude)?[=:]\s*(-?\d+\.\d+)', line, re.I)
                        lon_match = re.search(r'lon(?:gitude)?[=:]\s*(-?\d+\.\d+)', line, re.I)
                        if lat_match and lon_match:
                            lat = float(lat_match.group(1))
                            lon = float(lon_match.group(1))
                            return f"lat={lat},lon={lon}"
                        return line.strip()[:80]
            
            # 方法2: 使用 settings 获取最后已知位置
            success, stdout, _ = ADBHelper.execute_command(
                ['adb', '-s', self.current_device.serial, 'shell', 'settings', 'get', 'secure', 'location_providers_allowed'],
                timeout=3
            )
            if success and 'gps' in stdout.lower():
                # GPS已启用，尝试获取更精确位置
                pass
            
            # 方法3: 尝试获取基站位置（粗略）
            success, stdout, _ = ADBHelper.execute_command(
                ['adb', '-s', self.current_device.serial, 'shell', 'dumpsys', 'telephony.registry'],
                timeout=5
            )
            if success and stdout:
                lines = stdout.strip().split('\n')
                for line in lines:
                    if 'mCellIdentity' in line:
                        # 提取基站信息
                        return f"cell:{line.strip()[:60]}"
            
            return "位置信息暂时不可用"
        except Exception as e:
            return f"获取失败: {str(e)[:30]}"
    
    def _record_call_location(self, phone_number: str, call_index: int, total_calls: int):
        """记录通话位置（在电话接通时调用）"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        location_str = self._get_device_location()
        
        record = {
            'timestamp': timestamp,
            'phone_number': phone_number,
            'location': location_str,
            'call_index': call_index,
            'total_calls': total_calls
        }
        self.location_records.append(record)
        
        # 更新显示
        self._update_location_display()
        record_text = f"[{timestamp}] 拨打: {phone_number} | {location_str[:40]}\n"
        self.location_list.append(record_text)
        
        # 更新状态
        self.location_count_label.setText(f"通话记录: {len(self.location_records)}")
        self.location_status.setText("已记录")
        self.location_status.set_status("success")
    
    def _init_empty_map(self):
        """初始化空白地图"""
        # 创建以北京为中心的地图
        m = folium.Map(
            location=[39.9042, 116.4074],
            zoom_start=12,
            tiles='OpenStreetMap'
        )
        
        # 添加提示标记
        folium.Marker(
            [39.9042, 116.4074],
            popup='等待通话位置记录...',
            icon=folium.Icon(color='gray', icon='info-sign')
        ).add_to(m)
        
        # 保存并加载
        self._load_map_to_view(m)
    
    def _load_map_to_view(self, map_obj: folium.Map):
        """将 Folium 地图加载到 QWebEngineView"""
        import tempfile
        import os
        
        # 创建临时文件
        temp_dir = tempfile.gettempdir()
        map_path = os.path.join(temp_dir, 'phone_tester_map.html')
        map_obj.save(map_path)
        
        # 加载到视图
        self.map_view.load(QUrl.fromLocalFile(map_path))
    
    def _update_location_display(self):
        """更新位置地图显示 - 使用真实地图"""
        if not self.location_records:
            self._init_empty_map()
            return
        
        # 解析所有记录的位置坐标
        valid_coords = []
        for record in self.location_records:
            coords = self._parse_location_coords(record['location'])
            if coords:
                valid_coords.append({
                    'lat': coords[0],
                    'lng': coords[1],
                    'phone': record['phone_number'],
                    'time': record['timestamp'],
                    'index': record['call_index']
                })
        
        if not valid_coords:
            # 没有有效坐标，显示空白地图
            self._init_empty_map()
            return
        
        # 创建地图，中心点为第一个有效位置
        center_lat = valid_coords[0]['lat']
        center_lng = valid_coords[0]['lng']
        
        m = folium.Map(
            location=[center_lat, center_lng],
            zoom_start=15,
            tiles='OpenStreetMap'
        )
        
        # 添加标记点
        for i, coord in enumerate(valid_coords):
            # 根据索引选择颜色
            colors = ['red', 'blue', 'green', 'purple', 'orange', 'darkred', 'darkblue', 'darkgreen']
            color = colors[i % len(colors)]
            
            folium.Marker(
                [coord['lat'], coord['lng']],
                popup=f"通话 #{coord['index']}<br>号码: {coord['phone']}<br>时间: {coord['time']}",
                tooltip=f"通话 #{coord['index']}",
                icon=folium.Icon(color=color, icon='phone', prefix='fa')
            ).add_to(m)
        
        # 如果有多个点，添加连线
        if len(valid_coords) > 1:
            points = [[c['lat'], c['lng']] for c in valid_coords]
            folium.PolyLine(
                points,
                color='#2196F3',
                weight=3,
                opacity=0.7,
                dash_array='5, 10'
            ).add_to(m)
        
        # 适应边界
        if len(valid_coords) > 1:
            m.fit_bounds([[c['lat'], c['lng']] for c in valid_coords], padding=[50, 50])
        
        # 加载地图
        self._load_map_to_view(m)
    
    def _parse_location_coords(self, location_str: str) -> tuple:
        """从位置字符串解析经纬度"""
        import re
        
        # 尝试匹配各种经纬度格式
        # 格式1: lat=39.9, lon=116.4
        match = re.search(r'lat[=:]\s*(-?\d+\.?\d*)[\s,;]+lon[=:]\s*(-?\d+\.?\d*)', location_str, re.I)
        if match:
            return (float(match.group(1)), float(match.group(2)))
        
        # 格式2: latitude: 39.9, longitude: 116.4
        match = re.search(r'latitude[=:]\s*(-?\d+\.?\d*)[\s,;]+longitude[=:]\s*(-?\d+\.?\d*)', location_str, re.I)
        if match:
            return (float(match.group(1)), float(match.group(2)))
        
        # 格式3: 39.9042,116.4074 (纯数字对)
        match = re.search(r'(-?\d+\.\d+)[,\s]+(-?\d+\.\d+)', location_str)
        if match:
            lat, lng = float(match.group(1)), float(match.group(2))
            # 验证范围（中国大致范围）
            if 3 <= lat <= 54 and 73 <= lng <= 135:
                return (lat, lng)
        
        return None
    
    def _clear_location_records(self):
        """清空位置记录"""
        self.location_records.clear()
        self.location_list.clear()
        self._init_empty_map()
        self.location_count_label.setText("通话记录: 0")
        self.location_status.setText("未定位")
        self.location_status.set_status("default")
        self.current_location_display.setText(f"{ICONS['location']} 等待记录...")
    
    def _export_location_tracks(self):
        """导出位置轨迹"""
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        from datetime import datetime
        
        if not self.location_records:
            QMessageBox.information(self, "提示", "没有位置记录可导出")
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self, "导出位置轨迹",
            f"call_locations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            "Text Files (*.txt);;CSV Files (*.csv)"
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write("通话位置记录\n")
                    f.write("=" * 60 + "\n\n")
                    for record in self.location_records:
                        f.write(f"时间: {record['timestamp']}\n")
                        f.write(f"拨打号码: {record['phone_number']}\n")
                        f.write(f"位置: {record['location']}\n")
                        f.write(f"通话序号: {record['call_index']}/{record['total_calls']}\n")
                        f.write("-" * 60 + "\n\n")
                QMessageBox.information(self, "成功", f"已导出到:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")

    def _create_adb_tab(self) -> QWidget:
        """创建 ADB 终端 Tab"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)
        
        # 标题栏
        title_layout = QHBoxLayout()
        title_icon = QLabel("🖥️")
        title_icon.setStyleSheet("")
        title_layout.addWidget(title_icon)
        
        title_label = QLabel("ADB 命令终端")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        title_label.setStyleSheet(f"color: {COLORS['text_primary']};")
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        
        # 设备状态
        self.adb_tab_device_status = StatusBadge("未连接设备", "default")
        title_layout.addWidget(self.adb_tab_device_status)
        
        layout.addLayout(title_layout)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet(f"background-color: {COLORS['divider']}; max-height: 1px;")
        layout.addWidget(line)
        
        # 终端输出区域
        self.adb_output = QTextEdit()
        self.adb_output.setReadOnly(True)
        self.adb_output.setFont(QFont("Consolas", 11))
        self.adb_output.setPlaceholderText("ADB 命令输出将显示在这里...\n\n使用说明：\n1. 在下方输入命令（不需要输入 adb shell）\n2. 按回车或点击发送按钮执行\n3. 使用快捷按钮快速执行常用命令")
        self.adb_output.setStyleSheet(f"""
            QTextEdit {{
                background-color: #1e1e2e;
                color: #cdd6f4;
                border: none;
                border-radius: 8px;
                padding: 12px;
                selection-background-color: {COLORS['primary']}40;
            }}
            QScrollBar:vertical {{
                background-color: #313244;
                width: 12px;
                border-radius: 6px;
            }}
            QScrollBar::handle:vertical {{
                background-color: #585b70;
                border-radius: 6px;
                min-height: 20px;
            }}
            QScrollBar::handle:vertical:hover {{
                background-color: #6c7086;
            }}
        """)
        # 终端输出区域填充空间
        layout.addWidget(self.adb_output, 1)  # stretch factor = 1
        
        # 命令输入区域
        input_layout = QHBoxLayout()
        input_layout.setSpacing(8)
        
        # 前缀标签
        prefix = QLabel("shell>")
        prefix.setStyleSheet(f"""
            color: {COLORS['success']};
            font-family: 'Consolas', monospace;
            font-size: 13px;
            font-weight: bold;
            padding: 0 8px;
        """)
        input_layout.addWidget(prefix)
        
        # 命令输入框
        self.adb_cmd_input = QLineEdit()
        self.adb_cmd_input.setPlaceholderText("输入命令，例如: getprop ro.product.model")
        self.adb_cmd_input.setFont(QFont("Consolas", 12))
        self.adb_cmd_input.setStyleSheet(f"""
            QLineEdit {{
                padding: 10px 12px;
                border: 2px solid {COLORS['border']};
                border-radius: 6px;
                font-size: 13px;
                background-color: {COLORS['card_bg']};
                font-family: 'Consolas', monospace;
            }}
            QLineEdit:hover {{
                border-color: {COLORS['primary_light']};
            }}
            QLineEdit:focus {{
                border-color: {COLORS['primary']};
            }}
        """)
        self.adb_cmd_input.returnPressed.connect(self._send_adb_command)
        input_layout.addWidget(self.adb_cmd_input, 1)
        
        # 发送按钮
        self.adb_send_btn = QPushButton("▶ 执行")
        self.adb_send_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['success']};
                color: white;
                padding: 10px 20px;
                font-size: 13px;
                font-weight: bold;
                border-radius: 6px;
                border: none;
            }}
            QPushButton:hover {{
                background-color: {COLORS['success']}dd;
            }}
            QPushButton:pressed {{
                background-color: {COLORS['success']}bb;
            }}
            QPushButton:disabled {{
                background-color: {COLORS['divider']};
                color: {COLORS['text_secondary']};
            }}
        """)
        self.adb_send_btn.clicked.connect(self._send_adb_command)
        input_layout.addWidget(self.adb_send_btn)
        
        layout.addLayout(input_layout)
        
        # 快捷命令按钮 - 分组显示
        quick_cmds_group = QGroupBox("常用快捷命令")
        quick_cmds_group.setStyleSheet(f"""
            QGroupBox {{
                font-weight: bold;
                border: 1px solid {COLORS['border']};
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                font-size: 12px;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: {COLORS['text_secondary']};
            }}
        """)
        quick_cmds_layout = QVBoxLayout(quick_cmds_group)
        quick_cmds_layout.setSpacing(8)
        
        # 第1行：设备信息类
        row1 = QHBoxLayout()
        row1.setSpacing(8)
        quick_cmds_1 = [
            ("📱 设备信息", "getprop ro.product.model"),
            ("📶 信号状态", "dumpsys telephony.registry | grep -E 'mSignalStrength|mCallState'"),
            ("🔋 电池", "dumpsys battery"),
            ("💾 内存", "cat /proc/meminfo | head -5"),
            ("📋 进程", "ps -A | head -10"),
            ("🌐 网络", "ifconfig"),
        ]
        for label, cmd in quick_cmds_1:
            btn = QPushButton(label)
            btn.setStyleSheet(self._get_quick_btn_style())
            btn.clicked.connect(lambda checked, c=cmd: self._quick_adb_command(c))
            row1.addWidget(btn)
        row1.addStretch()
        quick_cmds_layout.addLayout(row1)
        
        # 第2行：系统操作类
        row2 = QHBoxLayout()
        row2.setSpacing(8)
        quick_cmds_2 = [
            ("🔄 重启设备", "reboot"),
            ("🔧 Recovery", "reboot recovery"),
            ("⚡ Fastboot", "reboot bootloader"),
            ("🛑 关机", "reboot -p"),
            ("📷 截图", "screencap -p /sdcard/screenshot.png"),
        ]
        for label, cmd in quick_cmds_2:
            btn = QPushButton(label)
            btn.setStyleSheet(self._get_quick_btn_style(COLORS['warning']))
            btn.clicked.connect(lambda checked, c=cmd: self._quick_adb_command(c))
            row2.addWidget(btn)
        row2.addStretch()
        quick_cmds_layout.addLayout(row2)
        
        # 第3行：调试类
        row3 = QHBoxLayout()
        row3.setSpacing(8)
        quick_cmds_3 = [
            ("🎯 当前Activity", "dumpsys activity activities | grep -E 'mResumedActivity'"),
            ("📜 日志", "logcat -d | tail -20"),
            ("📦 包列表", "pm list packages | head -10"),
            ("🎮 输入文字", "input text 'Hello'"),
            ("👆 点击屏幕", "input tap 500 1000"),
        ]
        for label, cmd in quick_cmds_3:
            btn = QPushButton(label)
            btn.setStyleSheet(self._get_quick_btn_style(COLORS['info']))
            btn.clicked.connect(lambda checked, c=cmd: self._quick_adb_command(c))
            row3.addWidget(btn)
        row3.addStretch()
        quick_cmds_layout.addLayout(row3)
        
        # 第4行：无线调试和高级功能
        row4 = QHBoxLayout()
        row4.setSpacing(8)
        quick_cmds_4 = [
            ("📡 无线调试", "tcpip 5555"),
            ("🔌 连接设备", "connect"),
            ("📤 推送到SD卡", "push /path/to/file /sdcard/"),
            ("📥 拉取文件", "pull /sdcard/file.txt ./"),
            ("🔍 屏幕分辨率", "wm size"),
        ]
        for label, cmd in quick_cmds_4:
            btn = QPushButton(label)
            btn.setStyleSheet(self._get_quick_btn_style(COLORS['success']))
            btn.clicked.connect(lambda checked, c=cmd: self._quick_adb_command(c))
            row4.addWidget(btn)
        row4.addStretch()
        quick_cmds_layout.addLayout(row4)
        
        layout.addWidget(quick_cmds_group)
        
        # 底部操作栏
        bottom_layout = QHBoxLayout()
        
        self.adb_clear_btn = QPushButton("🗑️ 清空输出")
        self.adb_clear_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {COLORS['text_secondary']};
                border: 1px solid {COLORS['border']};
                border-radius: 6px;
                padding: 6px 12px;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['warning']}20;
                color: {COLORS['warning']};
            }}
        """)
        self.adb_clear_btn.clicked.connect(self.adb_output.clear)
        bottom_layout.addWidget(self.adb_clear_btn)
        
        self.adb_save_btn = QPushButton("💾 保存输出")
        self.adb_save_btn.setStyleSheet(self.adb_clear_btn.styleSheet())
        self.adb_save_btn.clicked.connect(self._save_adb_output)
        bottom_layout.addWidget(self.adb_save_btn)
        
        bottom_layout.addStretch()
        
        # 命令历史
        history_label = QLabel("📜 历史:")
        history_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 12px;")
        bottom_layout.addWidget(history_label)
        
        self.adb_history_combo = QComboBox()
        self.adb_history_combo.setPlaceholderText("选择历史命令...")
        self.adb_history_combo.setMinimumWidth(150)
        self.adb_history_combo.setStyleSheet(f"""
            QComboBox {{
                padding: 6px 10px;
                border: 1px solid {COLORS['border']};
                border-radius: 6px;
                font-size: 12px;
            }}
        """)
        self.adb_history_combo.activated.connect(self._on_history_selected)
        bottom_layout.addWidget(self.adb_history_combo)
        
        layout.addLayout(bottom_layout)
        
        # 存储命令历史
        self.adb_command_history = []
        
        return tab
    
    def _send_adb_command(self):
        """发送 ADB 命令"""
        if not self.current_device:
            self._append_adb_output("❌ 错误: 请先连接设备！", "error")
            return
        
        cmd = self.adb_cmd_input.text().strip()
        if not cmd:
            return
        
        # 添加到历史
        if cmd not in self.adb_command_history:
            self.adb_command_history.insert(0, cmd)
            self.adb_history_combo.insertItem(0, cmd)
            if self.adb_history_combo.count() > 20:
                self.adb_history_combo.removeItem(20)
        
        # 显示执行的命令
        self._append_adb_output(f"$ {cmd}", "command")
        
        # 执行命令
        full_cmd = ['adb', '-s', self.current_device.serial, 'shell'] + cmd.split()
        success, stdout, stderr = ADBHelper.execute_command(full_cmd, timeout=30)
        
        if success:
            if stdout:
                self._append_adb_output(stdout, "output")
            else:
                self._append_adb_output("(命令执行成功，无输出)", "info")
        else:
            error_msg = stderr if stderr else "命令执行失败"
            self._append_adb_output(f"❌ {error_msg}", "error")
        
        # 清空输入框
        self.adb_cmd_input.clear()
    
    def _quick_adb_command(self, cmd: str):
        """执行快捷命令"""
        self.adb_cmd_input.setText(cmd)
        self._send_adb_command()
    
    def _on_history_selected(self, index: int):
        """选择历史命令"""
        if index >= 0:
            cmd = self.adb_history_combo.itemText(index)
            self.adb_cmd_input.setText(cmd)
            self.adb_cmd_input.setFocus()
    
    def _append_adb_output(self, text: str, style: str = "output"):
        """追加 ADB 输出到终端"""
        colors = {
            "command": "#89b4fa",
            "output": "#cdd6f4",
            "error": "#f38ba8",
            "info": "#fab387",
            "success": "#a6e3a1",
        }
        
        from PyQt6.QtGui import QTextCharFormat, QBrush, QColor
        
        cursor = self.adb_output.textCursor()
        cursor.movePosition(cursor.MoveOperation.End)
        
        format = QTextCharFormat()
        format.setForeground(QBrush(QColor(colors.get(style, colors["output"]))))
        
        if style == "command":
            from datetime import datetime
            timestamp = datetime.now().strftime("%H:%M:%S")
            cursor.insertText(f"\n[{timestamp}] ", format)
        
        cursor.insertText(text + "\n", format)
        
        # 自动滚动
        self.adb_output.verticalScrollBar().setValue(
            self.adb_output.verticalScrollBar().maximum()
        )
    
    def _get_quick_btn_style(self, hover_color=None) -> str:
        """获取快捷按钮样式"""
        if hover_color is None:
            hover_color = COLORS['primary']
        return f"""
            QPushButton {{
                background-color: {COLORS['background']};
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
                border-radius: 6px;
                padding: 6px 12px;
                font-size: 11px;
            }}
            QPushButton:hover {{
                background-color: {hover_color}20;
                border-color: {hover_color};
            }}
        """
    
    def _save_adb_output(self):
        """保存 ADB 输出到文件"""
        from PyQt6.QtWidgets import QFileDialog
        from datetime import datetime
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "保存 ADB 输出",
            f"adb_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            "Text Files (*.txt);;All Files (*.*)"
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(self.adb_output.toPlainText())
                self._append_adb_output(f"✅ 输出已保存到: {filename}", "success")
            except Exception as e:
                self._append_adb_output(f"❌ 保存失败: {str(e)}", "error")
    
    def _create_stat_card(self, icon: str, title: str, value: str, color: str) -> CardFrame:
        """创建统计卡片"""
        card = CardFrame()
        layout = QVBoxLayout(card)
        layout.setSpacing(8)
        layout.setContentsMargins(16, 16, 16, 16)
        
        # 图标和标题
        header = QLabel(f"{icon} {title}")
        header.setFont(QFont("Segoe UI", 12))
        header.setStyleSheet(f"color: {COLORS['text_secondary']};")
        layout.addWidget(header)
        
        # 数值
        value_label = QLabel(value)
        value_label.setFont(QFont("Segoe UI", 24, QFont.Weight.Bold))
        value_label.setStyleSheet(f"color: {color};")
        layout.addWidget(value_label)
        
        # 保存引用以便后续更新
        card.value_label = value_label
        
        return card
    
    def _toggle_auto_refresh(self, state):
        """切换自动检测状态"""
        if hasattr(self, 'device_check_timer'):
            if state == Qt.CheckState.Checked.value:
                self.device_check_timer.start(10000)
                self.log("已开启自动检测设备", "info")
            else:
                self.device_check_timer.stop()
                self.log("已关闭自动检测（手动点击刷新）", "info")
    
    def _set_adb_path_manual(self):
        """手动设置ADB路径"""
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        
        # 显示提示信息
        QMessageBox.information(
            self,
            "选择ADB可执行文件",
            "请找到并选择 adb.exe 文件\n\n"
            "常见位置：\n"
            "• Android Studio: C:\\Users\\<用户名>\\AppData\\Local\\Android\\Sdk\\platform-tools\\adb.exe\n"
            "• 独立安装: C:\\platform-tools\\adb.exe\n"
            "• 其他: 您下载的platform-tools文件夹中"
        )
        
        # 打开文件选择对话框
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择ADB可执行文件",
            "",
            "ADB Executable (adb.exe);;All Files (*.*)"
        )
        
        if file_path:
            if os.path.basename(file_path).lower() in ['adb.exe', 'adb']:
                if ADBHelper.set_adb_path(file_path):
                    self.log(f"✅ ADB路径已设置: {file_path}", "success")
                    self.status_bar.showMessage(f"{ICONS['success']} ADB路径已手动指定")
                    # 重置ADB可用性缓存，重新检测
                    ADBHelper._adb_available = None
                    self.refresh_devices()
                else:
                    QMessageBox.warning(self, "无效文件", "选择的文件无法执行，请确认是有效的adb.exe")
            else:
                QMessageBox.warning(self, "文件名错误", f"请选择 adb.exe，而不是 {os.path.basename(file_path)}")
    
    def start_device_check(self):
        """启动设备检测定时器 - 优化频率避免卡顿"""
        self.device_check_timer = QTimer(self)
        self.device_check_timer.timeout.connect(self.refresh_devices)
        self.device_check_timer.start(10000)  # 每10秒检测一次（降低频率避免卡顿）
        self.refresh_devices()
    
    def refresh_devices(self):
        """刷新设备列表 - 带状态变化检测"""
        # 首先检查ADB是否已安装
        if not ADBHelper.check_adb_installed():
            self.device_combo.clear()
            self.device_combo.addItem(f"{ICONS['warning']} 未找到ADB，点击手动指定")
            self.clear_device_info()
            self.status_bar.showMessage(f"{ICONS['error']} ADB未找到 - 请安装或手动指定路径")
            self.connection_status.setText("未找到ADB")
            self.connection_status.set_status("error")
            # 强制显示关键状态变化
            if self._last_connection_status != "no_adb":
                self._last_connection_status = "no_adb"
                self.log("未找到ADB工具，常见路径已搜索完毕", "error")
                self.log("请尝试：1. 点击刷新按钮右侧的'手动指定ADB' 2. 或安装Android Studio", "info")
            return
        
        devices = ADBHelper.get_devices()
        device_count = len(devices)
        
        current_serial = self.device_combo.currentData()
        self.device_combo.clear()
        
        if not devices:
            self.device_combo.addItem(f"{ICONS['disconnected']} 未检测到设备")
            self.clear_device_info()
            self.status_bar.showMessage(f"{ICONS['disconnected']} 未连接设备")
            self.connection_status.setText("未连接")
            self.connection_status.set_status("default")
            # 状态变化时强制显示日志
            if self._last_connection_status != "disconnected":
                self._last_connection_status = "disconnected"
                self._last_device_count = 0
                self.log("未检测到USB连接的设备", "warning")
            return
        
        # 设备已连接
        for serial in devices:
            self.device_combo.addItem(f"{ICONS['device']} {serial}", serial)
        
        # 恢复之前的选择
        if current_serial:
            index = self.device_combo.findData(current_serial)
            if index >= 0:
                self.device_combo.setCurrentIndex(index)
        
        # 如果只有一个设备，自动选择
        if len(devices) == 1:
            self.on_device_selected(0)
        
        # 状态变化时强制显示日志
        if self._last_connection_status != "connected" or self._last_device_count != device_count:
            if self._last_connection_status == "disconnected" or self._last_connection_status == "no_adb":
                # 从断开到连接的转变
                self.log(f"设备已连接，检测到 {device_count} 个设备", "success")
            elif self._last_device_count != device_count and self._last_device_count != -1:
                # 设备数量变化
                self.log(f"设备数量变化: {self._last_device_count} → {device_count}", "info")
            else:
                # 首次检测
                self.log(f"检测到 {device_count} 个设备", "success")
            
            self._last_connection_status = "connected"
            self._last_device_count = device_count
        
        self.connection_status.setText("已连接")
        self.connection_status.set_status("success")
    
    def on_device_selected(self, index: int):
        """设备选择改变时"""
        if index < 0:
            return
        
        serial = self.device_combo.currentData()
        if not serial or serial == "未检测到设备":
            return
        
        self.log(f"正在获取设备 {serial} 的信息...", "info")
        
        # 在新线程中获取设备信息
        def get_info():
            info = ADBHelper.get_device_info(serial)
            self.current_device = info
            self.update_device_display(info)
        
        thread = threading.Thread(target=get_info)
        thread.daemon = True
        thread.start()
    
    def update_device_display(self, info: DeviceInfo):
        """更新设备信息显示（支持双卡）"""
        # 基本信息
        self.device_name_label.setText(info.name)
        self.device_model_label.setText(info.model)
        self.android_version_label.setText(f"Android {info.android_version}" if info.android_version != "Unknown" else "未获取")
        
        # 更新卡一信息
        self._update_sim_display("sim1", info.sim1)
        # 更新卡二信息
        self._update_sim_display("sim2", info.sim2)
        
        # 更新 ADB 终端设备状态
        self.adb_device_status.setText("已连接")
        self.adb_device_status.set_status("success")
        self._append_adb_output(f"✅ 设备已连接: {info.name}", "success")
        
        self.log(f"设备信息已更新: {info.name}", "success")
    
    def _update_sim_display(self, sim_id: str, sim: SimInfo):
        """更新单个SIM卡的显示"""
        # 运营商
        operator_label = getattr(self, f"{sim_id}_operator_label")
        operator_label.setText(sim.operator if sim.operator != "Unknown" else "--")
        
        # 手机号
        number_label = getattr(self, f"{sim_id}_number_label")
        number_label.setText(sim.phone_number if sim.phone_number != "Unknown" else "--")
        
        # 网络类型
        network_label = getattr(self, f"{sim_id}_network_label")
        network_label.setText(sim.network_type if sim.network_type != "Unknown" else "--")
        
        # 状态标签
        status_badge = getattr(self, f"{sim_id}_status_badge")
        status_badge.setText(sim.state)
        if sim.state == "就绪":
            status_badge.set_status("success")
        elif sim.state == "未插入":
            status_badge.set_status("default")
        else:
            status_badge.set_status("warning")
        
        # 信号强度
        signal_bars = getattr(self, f"{sim_id}_signal_bars")
        signal_value = getattr(self, f"{sim_id}_signal_value")
        
        # 根据信号强度设置颜色
        signal_colors = {
            0: COLORS['divider'],
            1: COLORS['error'],
            2: COLORS['warning'],
            3: COLORS['info'],
            4: COLORS['success']
        }
        signal_texts = {
            0: "无信号",
            1: "弱",
            2: "一般",
            3: "良好",
            4: "强"
        }
        
        for i, bar in enumerate(signal_bars):
            if i < sim.signal_level:
                bar.setStyleSheet(f"""
                    QFrame {{
                        background-color: {signal_colors.get(sim.signal_level, COLORS['divider'])};
                        border-radius: 2px;
                    }}
                """)
            else:
                bar.setStyleSheet(f"""
                    QFrame {{
                        background-color: {COLORS['divider']};
                        border-radius: 2px;
                    }}
                """)
        
        signal_value.setText(signal_texts.get(sim.signal_level, "未知"))
    
    def clear_device_info(self):
        """清除设备信息显示"""
        self.current_device = None
        
        # 清除基本信息
        self.device_name_label.setText("未连接")
        self.device_model_label.setText("未连接")
        self.android_version_label.setText("未连接")
        
        # 清除双卡信息
        for sim_id in ["sim1", "sim2"]:
            # 运营商
            getattr(self, f"{sim_id}_operator_label").setText("--")
            # 手机号
            getattr(self, f"{sim_id}_number_label").setText("--")
            # 网络类型
            getattr(self, f"{sim_id}_network_label").setText("--")
            # 状态标签
            status_badge = getattr(self, f"{sim_id}_status_badge")
            status_badge.setText("未插入")
            status_badge.set_status("default")
            # 信号条
            signal_bars = getattr(self, f"{sim_id}_signal_bars")
            for bar in signal_bars:
                bar.setStyleSheet(f"""
                    QFrame {{
                        background-color: {COLORS['divider']};
                        border-radius: 2px;
                    }}
                """)
            # 信号值
            getattr(self, f"{sim_id}_signal_value").setText("无信号")
        
        # 清除 ADB 终端设备状态
        if hasattr(self, 'adb_device_status'):
            self.adb_device_status.setText("未连接")
            self.adb_device_status.set_status("default")
    
    def start_calling(self):
        """开始拨打 - 使用右侧策略配置"""
        # 检查设备
        if not self.current_device:
            QMessageBox.warning(self, f"{ICONS['warning']} 警告", "请先连接设备！")
            return
        
        # 检查是否有策略
        if not hasattr(self, 'strategies') or len(self.strategies) == 0:
            # 如果没有策略，切换到策略Tab提示用户
            self.tab_widget.setCurrentIndex(0)
            QMessageBox.warning(self, f"{ICONS['warning']} 警告", 
                "请先配置测试策略！\n\n请在右侧\"测试策略\"Tab中添加至少一个拨打策略。")
            return
        
        # 检查SIM卡状态
        if self.current_device.sim1.state != "就绪" and self.current_device.sim2.state != "就绪":
            reply = QMessageBox.question(
                self, f"{ICONS['warning']} 确认",
                "两张SIM卡都未就绪，是否继续？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return
        
        # 切换到日志Tab
        self.tab_widget.setCurrentIndex(1)
        
        # 开始执行策略
        self._execute_strategies()
    
    def _execute_strategies(self):
        """执行策略列表（支持电话和短信）"""
        if not self.strategies:
            return
        
        self.log("开始执行测试策略...", "info")
        self.log(f"共有 {len(self.strategies)} 个策略", "info")
        
        # 获取第一个策略的信息
        strategy = self.strategies[0]
        test_type = strategy.get('test_type', 'call')
        phone_number = strategy['target_num']
        
        # 提取主号码（如果有多个）
        if '/' in phone_number:
            phone_number = phone_number.split('/')[0].strip()
        if '(' in phone_number:
            phone_number = phone_number.split('(')[0].strip()
        
        # 根据测试类型执行不同操作
        if test_type == 'sms':
            self._execute_sms_strategy(strategy, phone_number)
        else:
            self._execute_call_strategy(strategy, phone_number)
    
    def _execute_call_strategy(self, strategy: dict, phone_number: str):
        """执行电话拨打策略"""
        duration = strategy['duration']
        count = strategy['count']
        
        self.log(f"【电话拨打】号码: {phone_number}, 时长: {duration}秒, 次数: {count}", "info")
        
        # 更新UI状态
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.test_btn.setEnabled(True)
        self.test_status_badge.setText("通话中")
        self.test_status_badge.set_status("info")
        
        # 清空之前的位置记录
        self._clear_location_records()
        
        # 创建CallWorker
        self.call_worker = CallWorker(
            self.current_device.serial,
            phone_number,
            duration,
            count
        )
        self.call_worker.log_signal.connect(self.on_log_received)
        self.call_worker.progress_signal.connect(self.on_progress)
        self.call_worker.status_signal.connect(self.on_status_update)
        self.call_worker.finished_signal.connect(self.on_calling_finished)
        # 连接位置记录信号
        self.call_worker.location_signal.connect(self._record_call_location)
        # 连接结果记录信号
        self.call_worker.result_signal.connect(self._add_result_record)
        self.call_worker.start()
    
    def _execute_sms_strategy(self, strategy: dict, phone_number: str):
        """执行短信发送策略"""
        count = strategy['count']
        sms_content = strategy.get('sms_content', '测试短信')
        local_sim = strategy['local_sim']
        sim_slot = 0 if "卡一" in local_sim else 1
        
        self.log(f"【短信发送】号码: {phone_number}, 内容: {sms_content[:20]}..., 次数: {count}", "info")
        
        # 更新UI状态
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.test_btn.setEnabled(True)
        self.test_status_badge.setText("发送短信中")
        self.test_status_badge.set_status("info")
        
        # 清空之前的位置记录
        self._clear_location_records()
        
        # 创建SMSWorker（使用QThread）
        self.sms_worker = SMSWorker(
            self.current_device.serial,
            phone_number,
            sms_content,
            sim_slot,
            count
        )
        self.sms_worker.log_signal.connect(self.on_log_received)
        self.sms_worker.progress_signal.connect(self.on_progress)
        self.sms_worker.status_signal.connect(self.on_status_update)
        self.sms_worker.finished_signal.connect(self.on_sms_finished)
        self.sms_worker.result_signal.connect(self._add_sms_result_record)
        self.sms_worker.start()
    
    def stop_calling(self):
        """停止拨打"""
        if self.call_worker and self.call_worker.isRunning():
            self.call_worker.stop()
            self.log("正在停止拨打任务...", "warning")
    
    def on_log_received(self, message: str, log_type: str):
        """接收到日志消息"""
        self.log(message, log_type)
    
    def on_progress(self, current: int, total: int):
        """进度更新"""
        self.progress_bar.setValue(current)
        self.stats_label.setText(f"{ICONS['call']} {current}/{total} 次")
    
    def on_status_update(self, status: str):
        """状态更新"""
        self.status_bar.showMessage(f"{ICONS['info']} {status}")
    
    def on_calling_finished(self):
        """拨打完成"""
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.device_combo.setEnabled(True)
        self.status_bar.showMessage(f"{ICONS['success']} 拨打测试完成")
        self.progress_bar.hide()
        self.stats_label.setText(f"{ICONS['check']} 完成")
        self.test_status_badge.setText("完成")
        self.test_status_badge.set_status("success")
    
    def on_sms_finished(self):
        """短信发送完成"""
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.device_combo.setEnabled(True)
        self.status_bar.showMessage(f"{ICONS['success']} 短信发送测试完成")
        self.progress_bar.hide()
        self.stats_label.setText(f"{ICONS['check']} 完成")
        self.test_status_badge.setText("完成")
        self.test_status_badge.set_status("success")
    
    def _add_sms_result_record(self, result_dict):
        """添加短信测试结果到结果表格"""
        # 复用电话测试的结果添加方法，但调整显示
        self._add_result_record(result_dict)
    
    def log(self, message: str, log_type: str = "info"):
        """添加日志 - 智能去重和关键事件高亮"""
        from datetime import datetime
        import time
        
        now = time.time()
        timestamp = datetime.now().strftime("%H:%M:%S")
        
        # 判断是否为关键事件
        is_critical = any(event in message for event in self._critical_events)
        
        # 重复消息检测（非关键事件才进行去重）
        if not is_critical:
            # 提取消息核心内容（去除数字变化部分）
            message_key = self._extract_message_key(message)
            last_shown = self._last_log_message.get(message_key, 0)
            
            if now - last_shown < self._log_dedup_interval:
                # 跳过显示，但记录到计数
                self._last_log_message[f"{message_key}_count"] = \
                    self._last_log_message.get(f"{message_key}_count", 0) + 1
                return
            
            # 检查是否有被跳过的重复消息需要补报
            skipped_count = self._last_log_message.get(f"{message_key}_count", 0)
            if skipped_count > 0:
                # 先显示被跳过的消息统计
                skip_msg = f"[{timestamp}] {ICONS['info']} ... ({skipped_count} 条相似消息已省略)\n"
                cursor = self.log_text.textCursor()
                cursor.movePosition(cursor.MoveOperation.End)
                fmt = QTextCharFormat()
                fmt.setForeground(QBrush(QColor("#6c7086")))  # 灰色
                fmt.setFontItalic(True)
                cursor.insertText(skip_msg, fmt)
                self._last_log_message[f"{message_key}_count"] = 0
            
            # 更新最后显示时间
            self._last_log_message[message_key] = now
        
        # 图标映射
        icon_map = {
            "info": ICONS['info'],
            "success": ICONS['success'],
            "error": ICONS['error'],
            "warning": ICONS['warning'],
        }
        icon = icon_map.get(log_type, ICONS['info'])
        
        # 构建日志条目
        if is_critical:
            # 关键事件使用更醒目的格式
            log_entry = f"[{timestamp}] {icon} ▶ {message}\n"
        else:
            log_entry = f"[{timestamp}] {icon} {message}\n"
        
        # 设置文本颜色
        color_map = {
            "info": QColor("#89b4fa"),      # 浅蓝
            "success": QColor("#a6e3a1"),   # 浅绿
            "error": QColor("#f38ba8"),     # 粉红
            "warning": QColor("#fab387"),   # 橙黄
        }
        color = color_map.get(log_type, QColor("#cdd6f4"))
        
        # 关键事件使用更亮的颜色
        if is_critical:
            if log_type == "info":
                color = QColor("#74c7ec")   # 更亮的蓝色
            elif log_type == "success":
                color = QColor("#94e2d5")   # 更亮的青色
        
        # 插入带颜色的文本
        cursor = self.log_text.textCursor()
        format = QTextCharFormat()
        format.setForeground(QBrush(color))
        
        # 关键事件加粗显示
        if is_critical:
            format.setFontWeight(QFont.Weight.Bold)
        
        cursor.movePosition(cursor.MoveOperation.End)
        cursor.insertText(log_entry, format)
        
        # 自动滚动到底部
        self.log_text.verticalScrollBar().setValue(
            self.log_text.verticalScrollBar().maximum()
        )
    
    def _extract_message_key(self, message: str) -> str:
        """提取消息核心内容用于去重判断（去除数字和时间等变化部分）"""
        import re
        # 替换数字为占位符
        key = re.sub(r'\d+', '#', message)
        # 替换时间格式
        key = re.sub(r'\d{1,2}:\d{2}:\d{2}', '[TIME]', key)
        # 替换括号内的计数
        key = re.sub(r'\([^)]*\d[^)]*\)', '(#)', key)
        return key
    
    def clear_log(self):
        """清空日志"""
        self.log_text.clear()
        self.log("日志已清空", "info")
    
    def save_log(self):
        """保存日志"""
        from PyQt6.QtWidgets import QFileDialog
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            f"{ICONS['save']} 保存日志",
            f"call_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            "Text Files (*.txt);;All Files (*.*)"
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(self.log_text.toPlainText())
                self.log(f"日志已保存到: {filename}", "success")
            except Exception as e:
                QMessageBox.critical(self, f"{ICONS['error']} 错误", f"保存日志失败: {str(e)}")
    
    def _on_splitter_moved(self, pos: int, index: int):
        """分割器移动时的回调 - 在状态栏显示当前尺寸"""
        sizes = self.main_splitter.sizes()
        left_width = sizes[0]
        right_width = sizes[1]
        total_width = left_width + right_width
        left_percent = (left_width / total_width) * 100 if total_width > 0 else 0
        
        self.status_bar.showMessage(
            f"{ICONS['resize']} 左侧: {left_width}px ({left_percent:.0f}%) | "
            f"右侧: {right_width}px ({100-left_percent:.0f}%) - 拖动分割线调整布局",
            3000  # 显示3秒
        )
    
    def reset_splitter(self):
        """重置分割器为默认比例"""
        if hasattr(self, 'main_splitter'):
            total_width = sum(self.main_splitter.sizes())
            if total_width > 0:
                # 默认比例 1:2 (左:右)
                left_width = int(total_width * 0.32)
                right_width = total_width - left_width
                self.main_splitter.setSizes([left_width, right_width])
                self.log("面板布局已重置为默认比例", "info")
    
    def closeEvent(self, event):
        """窗口关闭事件"""
        if self.call_worker and self.call_worker.isRunning():
            reply = QMessageBox.question(
                self, f"{ICONS['warning']} 确认",
                "正在执行拨打任务，是否确认退出？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.call_worker.stop()
                self.call_worker.wait(3000)
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()


def main():
    app = QApplication(sys.argv)
    
    # 根据操作系统设置风格
    if platform.system() == 'Windows':
        # Windows 使用原生风格
        app.setStyle('Windows')
        # 设置 Windows 字体
        font = QFont("Microsoft YaHei", 9)
        app.setFont(font)
    else:
        # macOS/Linux 使用 Fusion 风格
        app.setStyle('Fusion')
    
    # 设置应用程序全局样式 - Windows 原生风格
    font_family = get_font_family()
    app.setStyleSheet(f"""
        QMainWindow {{
            background-color: {COLORS['background']};
        }}
        QWidget {{
            font-family: {font_family};
            font-size: 9pt;
        }}
        QLabel {{
            color: {COLORS['text_primary']};
        }}
        QGroupBox {{
            font-weight: 600;
            border: 1px solid {COLORS['border']};
            border-radius: 4px;
            margin-top: 8px;
            padding: 12px;
            background-color: {COLORS['card_bg']};
        }}
        QGroupBox::title {{
            subcontrol-origin: margin;
            left: 8px;
            padding: 0 6px;
            color: {COLORS['text_primary']};
            font-size: 9pt;
        }}
        QPushButton {{
            font-family: {font_family};
            font-size: 9pt;
            padding: 5px 12px;
            border: 1px solid {COLORS['border']};
            border-radius: 2px;
            background-color: {COLORS['card_bg']};
        }}
        QPushButton:hover {{
            background-color: {COLORS['primary_light']};
            border-color: {COLORS['primary']};
        }}
        QPushButton:pressed {{
            background-color: {COLORS['primary']};
            color: white;
        }}
        QComboBox {{
            font-family: {font_family};
            font-size: 9pt;
            padding: 4px 8px;
            border: 1px solid {COLORS['border']};
            border-radius: 2px;
            background-color: {COLORS['card_bg']};
        }}
        QLineEdit {{
            font-family: {font_family};
            font-size: 9pt;
            padding: 4px 8px;
            border: 1px solid {COLORS['border']};
            border-radius: 2px;
        }}
        QSpinBox {{
            font-family: {font_family};
            font-size: 9pt;
            padding: 4px;
            border: 1px solid {COLORS['border']};
            border-radius: 2px;
        }}
        QTextEdit {{
            font-family: {get_mono_font()};
            font-size: 9pt;
            border: 1px solid {COLORS['border']};
            border-radius: 2px;
        }}
        QTabWidget::pane {{
            border: 1px solid {COLORS['border']};
            background-color: {COLORS['card_bg']};
        }}
        QTabBar::tab {{
            font-family: {font_family};
            font-size: 9pt;
            padding: 6px 12px;
            margin-right: 2px;
            border: 1px solid {COLORS['border']};
            border-bottom: none;
            border-top-left-radius: 2px;
            border-top-right-radius: 2px;
            background-color: {COLORS['background']};
        }}
        QTabBar::tab:selected {{
            background-color: {COLORS['card_bg']};
            border-bottom: 2px solid {COLORS['primary']};
        }}
        QFormLayout QLabel {{
            color: {COLORS['text_secondary']};
        }}
        QMessageBox {{
            background-color: {COLORS['card_bg']};
        }}
        QDialog {{
            background-color: {COLORS['card_bg']};
        }}
    """)
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
